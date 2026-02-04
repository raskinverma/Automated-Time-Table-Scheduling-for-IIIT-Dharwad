import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
RANDOM_SEED = 42
random.seed(RANDOM_SEED)

class Course:
    def __init__(self, row):
        self.code = str(row["Course_Code"]).strip()
        self.basket = int(row.get("basket", 0))
        self.title = str(row.get("Course_Title", self.code)).strip()
        self.faculty = str(row.get("Faculty", "")).strip()
        self.ltp = str(row["L-T-P-S-C"]).strip()
        self.semester_half = str(row.get("Semester_Half", "0")).strip()
        self.is_elective = str(row.get("Elective", 0)).strip() == "1"
        try:
            self.L, self.T, self.P, self.S, self.C = map(int, self.ltp.split("-"))
        except Exception:
            self.L, self.T, self.P = 0, 0, 0


class Scheduler:
    def __init__(self, slots_file, courses_file, rooms_file, global_room_usage, global_elective_slots=None):
        df = pd.read_csv(slots_file)
        self.slots = [f"{row['Start_Time'].strip()}-{row['End_Time'].strip()}" for _, row in df.iterrows()]
        self.slot_durations = {s: self._slot_duration(s) for s in self.slots}

        courses_df = pd.read_csv(courses_file)
        self.courses = [Course(row) for _, row in courses_df.iterrows()]

        rooms_df = pd.read_csv(rooms_file)
        self.classrooms = []
        self.labs = []
        self.all_rooms = []
        for _, row in rooms_df.iterrows():
            room_id = str(row["Room_ID"]).strip()
            self.all_rooms.append(room_id)
            if room_id.upper().startswith("L"):
                self.labs.append(room_id)
            elif room_id.upper().startswith("C"):
                self.classrooms.append(room_id)
            else:
                continue

        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.excluded_slots = ["07:30-09:00", "13:15-14:00"]
        self.MAX_ATTEMPTS = 2000
        self.unscheduled_courses = []
        self.course_room_map = {}
        self.global_room_usage = global_room_usage
        self.scheduled_entries = []
        self.electives_by_sheet = {}
        self.elective_room_assignment = {}
        self.break_length_slots = 1
        self.global_elective_slots = global_elective_slots if global_elective_slots is not None else {}


 
    def _slot_duration(self, slot):
        start, end = slot.split("-")
        h1, m1 = map(int, start.split(":"))
        h2, m2 = map(int, end.split(":"))
        return (h2 + m2 / 60) - (h1 + m1 / 60)

    def _get_free_blocks(self, timetable, day):
        free_blocks, block = [], []
        for slot in self.slots:
            if timetable.at[day, slot] == "" and slot not in self.excluded_slots:
                block.append(slot)
            else:
                if block:
                    free_blocks.append(block)
                    block = []
        if block:
            free_blocks.append(block)
        return free_blocks

    def _allocate_session(
        self,
        timetable,
        lecturer_busy,
        labs_scheduled,
        day,
        faculty,
        code,
        duration_hours,
        session_type="L",
        is_elective=False,
        sheet_name=None,
        force_slots=None,
        min_rooms_needed=1
    ):
       
        for entry in self.scheduled_entries:
            if entry["day"] == day and entry["code"] == code and entry["sheet"] == sheet_name:
                return None
       
        
        if session_type == "P" and labs_scheduled[day]:
            return None

        # Try to find valid slots
        valid_slots_found = None
        room_to_use = ""

        if force_slots:
            # FORCE MODE: Check if provided slots are free in local timetable
            conflict = False
            for s in force_slots:
                if timetable.at[day, s] != "":
                    conflict = True
                    break
            
            if not conflict:
                # Check faculty busy
                if faculty and any(faculty in lecturer_busy[day][s] for s in force_slots):
                    conflict = True
                
            if not conflict:
                valid_slots_found = force_slots
                if is_elective and min_rooms_needed > 1:
                     # Verify capacity even for forced slots (though typically they are pre-validated)
                     if any(len(self.all_rooms) - len(self.global_room_usage.get(day, {}).get(s, [])) < min_rooms_needed for s in force_slots):
                         valid_slots_found = None

        else:
            # SEARCH MODE: Find all candidate sub-blocks that fit the duration
            free_blocks = self._get_free_blocks(timetable, day)
            candidates = []

            for block in free_blocks:
                # Sliding window over the block
                for i in range(len(block)):
                    dur_accum = 0
                    current_slots = []
                    for j in range(i, len(block)):
                        s = block[j]
                        current_slots.append(s)
                        dur_accum += self.slot_durations[s]
                        
                        if dur_accum >= duration_hours:
                            # Found a valid sub-block
                            waste = dur_accum - duration_hours
                            candidates.append({
                                'slots': current_slots,
                                'waste': waste,
                                'start_idx': i # mainly for stability if needed
                            })
                            # Once we reach required duration, we can stop extending strictly for "minimum fit"
                            # But if the next slot creates a "better" fit (unlikely if duration increases), we'd continue.
                            # Since slot durations are positive, adding more slots only increases waste. 
                            # So we stop this inner extension loop.
                            break 
            
            # Sort candidates by waste (ascending) to prefer tight fits
            candidates.sort(key=lambda x: x['waste'])

            for cand in candidates:
                current_slots = cand['slots']
                
                # Check faculty availability
                if faculty and any(faculty in lecturer_busy[day][s] for s in current_slots):
                    continue 

                # Check room capacity for electives
                if is_elective and min_rooms_needed > 1:
                     if any(len(self.all_rooms) - len(self.global_room_usage.get(day, {}).get(s, [])) < min_rooms_needed for s in current_slots):
                         continue

                # Check Room Availability
                chosen_room = ""
                if not is_elective:
                    mapped = self.course_room_map.get(code)
                    if mapped:
                        # Check if mapped room is free
                        if (session_type == "P" and mapped.upper().startswith("L")) or (session_type != "P" and not mapped.upper().startswith("L")):
                            if all(mapped not in self.global_room_usage.get(day, {}).get(s, []) for s in current_slots):
                                chosen_room = mapped
                            else:
                                mapped = None # Mapped room busy, try others
                        else:
                            mapped = None
                    
                    if not chosen_room:
                        possible_rooms = self.labs if session_type == "P" else self.classrooms
                        random.shuffle(possible_rooms)
                        available_rooms = [
                            r for r in possible_rooms
                            if all(r not in self.global_room_usage.get(day, {}).get(s, []) for s in current_slots)
                        ]
                        if not available_rooms:
                             continue
                        chosen_room = available_rooms[0]

                # Found a valid candidate!
                valid_slots_found = current_slots
                room_to_use = chosen_room
                break
        
        if not valid_slots_found:
            return None

        # Apply allocation
        slots_to_use = valid_slots_found
        
        if not is_elective:
            if not self.course_room_map.get(code) and room_to_use:
                 self.course_room_map[code] = room_to_use
            
            if room_to_use:
                for s in slots_to_use:
                     self.global_room_usage.setdefault(day, {}).setdefault(s, []).append(room_to_use)

        for i, s in enumerate(slots_to_use):
            if session_type == "L":
                display_text = f"{code} ({room_to_use})" if (room_to_use and not is_elective) else code
            elif session_type == "T":
                display_text = f"{code}T ({room_to_use})" if (room_to_use and not is_elective) else f"{code}T"
            elif session_type == "P":
                suffix = f" (Lab-{room_to_use})" if room_to_use and not is_elective else " (Lab)"
                display_text = f"{code}{suffix}"
            else:
                display_text = code

            timetable.at[day, s] = display_text

            self.scheduled_entries.append(
                {
                    "sheet": sheet_name,
                    "day": day,
                    "slot": s,
                    "code": code,
                    "display": display_text,
                    "faculty": faculty,
                    "room": room_to_use,
                }
            )

            # Mark gap slot if applies
            if i < len(slots_to_use) - 1:
                idx = self.slots.index(s)
                if idx + 1 < len(self.slots):
                    gap_slot = self.slots[idx + 1]
                    if timetable.at[day, gap_slot] == "" and self.slot_durations[gap_slot] == 0.25:
                        timetable.at[day, gap_slot] = "FREE"

        if faculty:
            for s in slots_to_use:
                lecturer_busy[day][s].append(faculty)

        if session_type == "P":
            labs_scheduled[day] = True
        
        # Add break
        last_slot = slots_to_use[-1]
        idx = self.slots.index(last_slot)
        for extra in range(1, self.break_length_slots + 1):
            if idx + extra < len(self.slots):
                next_slot = self.slots[idx + extra]
                if timetable.at[day, next_slot] == "":
                    timetable.at[day, next_slot] = "BREAK"
                    if faculty:
                        lecturer_busy[day][next_slot].append(faculty)
                    if not is_elective and room_to_use:
                        self.global_room_usage.setdefault(day, {}).setdefault(next_slot, []).append(room_to_use)

        return slots_to_use


    def generate_timetable(self, courses_to_allocate, writer, sheet_name):
        timetable = pd.DataFrame("", index=self.days, columns=self.slots)
        lecturer_busy = {day: {slot: [] for slot in self.slots} for day in self.days}
        labs_scheduled = {day: False for day in self.days}
        self.course_room_map = {}

        electives = [c for c in courses_to_allocate if c.is_elective]
        non_electives = [c for c in courses_to_allocate if not c.is_elective]

        baskets = {}
        for e in electives:
            baskets.setdefault(e.basket, []).append(e)

        basket_sizes = {b: len(g) for b, g in baskets.items()}

        chosen_electives = []

        self.electives_by_sheet[sheet_name] = chosen_electives

        elective_placeholders = []

        for b in sorted(baskets.keys()):
            if b == 0:
                continue
            group = baskets[b]
            chosen = random.choice(group)
            chosen_electives.append((b, chosen))

            elective_course = Course(
                {
                    "Course_Code": f"Elective_{b}",
                    "Course_Title": chosen.title,
                    "Faculty": chosen.faculty,
                    "L-T-P-S-C": chosen.ltp,
                    "Semester_Half": chosen.semester_half,
                    "Elective": 0,
                }
            )
            elective_placeholders.append(elective_course)

        self.electives_by_sheet[sheet_name] = chosen_electives

        # Sort non-electives to prioritize harder tasks (Labs > Lectures > Tutorials)
        non_electives.sort(key=lambda c: (c.P, c.L + c.T + c.S), reverse=True)
        
        # Prioritize electives:
        all_courses = elective_placeholders + non_electives

        for course in all_courses:
            faculty, code, is_elective = course.faculty, course.code, course.code.startswith("Elective_")
            basket_id = int(code.split("_")[1]) if is_elective else None
            min_rooms = basket_sizes.get(basket_id, 1) if is_elective else 1

            # --- Lecture Scheduling ---
            # Prepare to schedule. If elective and global slots exist, use them.
            
            forced_allocations_L = []
            if is_elective:
                key = (basket_id, "L")
                if key in self.global_elective_slots:
                    forced_allocations_L = self.global_elective_slots[key]

            remaining = course.L
            
            if forced_allocations_L:
                # Deterministic scheduling for this elective
                for alloc in forced_allocations_L:
                    day = alloc['day']
                    slots = alloc['slots']
                    duration = sum(self.slot_durations[s] for s in slots)
                    # Try to allocate forced slots
                    res = self._allocate_session(
                         timetable, lecturer_busy, labs_scheduled, day, faculty, code, duration, "L", is_elective, sheet_name, force_slots=slots, min_rooms_needed=min_rooms
                    )
                    if res:
                        remaining -= duration
            
            # Standard stochastic scheduling (runs if not fully scheduled by force)
            attempts = 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_to_try = self.days.copy()
                random.shuffle(days_to_try)
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                        continue
                    alloc_dur = min(1.5, remaining)
                    
                    allocated_slots = self._allocate_session(
                        timetable, lecturer_busy, labs_scheduled, day, faculty, code, alloc_dur, "L", is_elective, sheet_name, min_rooms_needed=min_rooms
                    )
                    if allocated_slots:
                        remaining -= alloc_dur
                        if is_elective:
                            self.global_elective_slots.setdefault((basket_id, "L"), []).append({
                                'day': day,
                                'slots': allocated_slots
                            })
                        break

            if remaining > 0:
                self.unscheduled_courses.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Lecture",
                    "remaining_hours": remaining,
                    "semester_half": course.semester_half
                })

            # --- Tutorial Scheduling ---
            forced_allocations_T = []
            if is_elective:
                key = (basket_id, "T")
                if key in self.global_elective_slots:
                    forced_allocations_T = self.global_elective_slots[key]

            remaining = course.T
            if forced_allocations_T:
                 for alloc in forced_allocations_T:
                    day = alloc['day']
                    slots = alloc['slots']
                    res = self._allocate_session(
                         timetable, lecturer_busy, labs_scheduled, day, faculty, code, 1, "T", is_elective, sheet_name, force_slots=slots, min_rooms_needed=min_rooms
                    )
                    if res:
                        remaining -= 1

            attempts = 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_to_try = self.days.copy()
                random.shuffle(days_to_try)
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                        continue
                    
                    allocated_slots = self._allocate_session(
                        timetable, lecturer_busy, labs_scheduled, day, faculty, code, 1, "T", is_elective, sheet_name, min_rooms_needed=min_rooms
                    )
                    if allocated_slots:
                        remaining -= 1
                        if is_elective:
                            self.global_elective_slots.setdefault((basket_id, "T"), []).append({
                                'day': day,
                                'slots': allocated_slots
                            })
                        break

            if remaining > 0:
                self.unscheduled_courses.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Tutorial",
                    "remaining_hours": remaining,
                    "semester_half": course.semester_half
                })

            # --- Lab Scheduling ---
            forced_allocations_P = []
            if is_elective:
                key = (basket_id, "P")
                if key in self.global_elective_slots:
                    forced_allocations_P = self.global_elective_slots[key]

            remaining = course.P
            if forced_allocations_P:
                 for alloc in forced_allocations_P:
                    day = alloc['day']
                    slots = alloc['slots']
                    duration = sum(self.slot_durations[s] for s in slots)
                    res = self._allocate_session(
                         timetable, lecturer_busy, labs_scheduled, day, faculty, code, duration, "P", is_elective, sheet_name, force_slots=slots, min_rooms_needed=min_rooms
                    )
                    if res:
                        remaining -= duration
            
            attempts = 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_without_labs = [d for d in self.days if not labs_scheduled[d]]
                days_to_try = days_without_labs.copy()
                random.shuffle(days_to_try)
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                        continue
                    alloc_dur = min(2, remaining) if remaining >= 2 else remaining
                    
                    allocated_slots = self._allocate_session(
                        timetable, lecturer_busy, labs_scheduled, day, faculty, code, alloc_dur, "P", is_elective, sheet_name, min_rooms_needed=min_rooms
                    )
                    if allocated_slots:
                        remaining -= alloc_dur
                        if is_elective:
                            self.global_elective_slots.setdefault((basket_id, "P"), []).append({
                                'day': day,
                                'slots': allocated_slots
                            })
                        break

            if remaining > 0:
                self.unscheduled_courses.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Lab",
                    "remaining_hours": remaining,
                    "semester_half": course.semester_half
                })

        for day in self.days:
            for slot in self.excluded_slots:
                if slot in timetable.columns:
                    timetable.at[day, slot] = ""

        timetable.to_excel(writer, sheet_name=sheet_name, index=True)
        print(f"Saved timetable to sheet '{sheet_name}'")


    def _compute_elective_room_assignments_legally(self, sheet_name):
        electives_representatives = self.electives_by_sheet.get(sheet_name, [])
        if not electives_representatives:
            self.elective_room_assignment[sheet_name] = {}
            return

        active_baskets = set(b for b, _ in electives_representatives)
        assigned = {}
        
        # Local usage tracker for electives in this sheet: day -> slot -> set of rooms used
        local_room_usage = {} 

        candidate_rooms = list(self.classrooms) + list(self.labs)

        for basket in sorted(active_baskets):
            basket_code = f"Elective_{basket}"
            # Get slots for this basket
            lecture_slots = []
            lab_slots = []
            for ent in self.scheduled_entries:
                if ent["sheet"] == sheet_name and ent["code"] == basket_code:
                    if "(Lab" in ent["display"]: 
                         lab_slots.append((ent["day"], ent["slot"]))
                    else:
                         lecture_slots.append((ent["day"], ent["slot"]))
            
            # Prioritize lecture slots for room assignment (Classrooms). 
            # If no lecture slots (pure lab), use lab slots.
            # This prevents Lab slots (which happen in Labs) from blocking Classrooms for the Lecture component.
            raw_slots = lecture_slots if lecture_slots else lab_slots
            basket_slots = sorted(list(set(raw_slots)))
            
            # Even if no slots found (e.g. unscheduled), we proceed (will act as 0 duration) 
            # but room assignment requires slots to check conflicts. 
            # If no slots, room assignment is arbitrary (free).
            
            basket_electives = [c for c in self.courses if c.is_elective and c.basket == basket]
            
            for elective in basket_electives:
                key = f"Elective_{basket}||{elective.title}"
                
                # OPTIMIZATION: Try to find ONE stable room for all slots first
                stable_room = None
                
                # Check all rooms for stability
                # Preference order: Classrooms first if this course actually has lectures (to avoid putting lectures in labs ideally)
                
                search_order = sorted(self.classrooms)
                if not lecture_slots: # If pure lab course, prefer labs
                     search_order = sorted(self.labs) + sorted(self.classrooms)
                
                for r in search_order: 
                     is_universally_free = True
                     for day, slot in basket_slots:
                         # Global check
                         if r in self.global_room_usage.get(day, {}).get(slot, []):
                             is_universally_free = False
                             break
                         # Local check
                         if r in local_room_usage.get(day, {}).get(slot, set()):
                             is_universally_free = False
                             break
                     
                     if is_universally_free:
                         stable_room = r
                         break
                
                used_rooms_for_this_elective = []
                
                if stable_room:
                    # Perfect! We found one room for everything.
                    for day, slot in basket_slots:
                        local_room_usage.setdefault(day, {}).setdefault(slot, set()).add(stable_room)
                    used_rooms_for_this_elective.append(stable_room)
                else:
                    # Fallback: Assign room PER SLOT
                    room_assignments = [] # List of (day, room)
                    
                    last_used_room = None
                    
                    for day, slot in basket_slots:
                        # Reuse last room if possible
                        candidates_ordered = []
                        if last_used_room:
                            candidates_ordered.append(last_used_room)
                            candidates_ordered.extend([r for r in candidate_rooms if r != last_used_room])
                        else:
                            candidates_ordered = candidate_rooms

                        chosen_for_slot = None
                        
                        for r in candidates_ordered:
                            if r in self.global_room_usage.get(day, {}).get(slot, []): continue
                            if r in local_room_usage.get(day, {}).get(slot, set()): continue
                            chosen_for_slot = r
                            break
                        
                        if not chosen_for_slot:
                             for r in sorted(self.all_rooms):
                                 if r not in self.global_room_usage.get(day, {}).get(slot, []):
                                     chosen_for_slot = r
                                     break
                        
                        if chosen_for_slot:
                            local_room_usage.setdefault(day, {}).setdefault(slot, set()).add(chosen_for_slot)
                            room_assignments.append((day, chosen_for_slot))
                            last_used_room = chosen_for_slot

                # Format the room string
                if stable_room:
                    assigned[key] = stable_room
                else:
                    # Group by room -> days
                    room_to_days = {}
                    for day, room in room_assignments:
                        room_to_days.setdefault(room, set()).add(day)
                    
                    # Sort rooms by first appearance day (approx) or name
                    # Let's sort by room name for consistency or by schedule order? 
                    # Schedule order is better but complex. Room name is fine.
                    
                    parts = []
                    # Pre-defined abbreviations
                    day_abbr = {"Monday": "Mon", "Tuesday": "Tue", "Wednesday": "Wed", "Thursday": "Thu", "Friday": "Fri"}
                    
                    # Sort logic: sort by first day index found for that room
                    def sort_key(item):
                        r, days = item
                        # Min index of days
                        indices = [self.days.index(d) for d in days if d in self.days]
                        return min(indices) if indices else 99
                    
                    sorted_items = sorted(room_to_days.items(), key=sort_key)
                    
                    for room, days in sorted_items:
                        # Sort days
                        sorted_days = sorted(list(days), key=lambda d: self.days.index(d) if d in self.days else 99)
                        d_strs = [day_abbr.get(d, d[:3]) for d in sorted_days]
                        parts.append(f"{room} ({','.join(d_strs)})")
                    
                    assigned[key] = ", ".join(parts)

        self.elective_room_assignment[sheet_name] = assigned

    def format_student_timetable_with_legend(self, filename):
        wb = load_workbook(filename)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        color_map = {}
        palette = ["FFC7CE", "C6EFCE", "FFEB9C", "BDD7EE", "D9EAD3", "F4CCCC",
                "D9D2E9", "FCE5CD", "C9DAF8", "EAD1DC"]
        color_index = 0

        for sheet_name in wb.sheetnames:
            self._compute_elective_room_assignments_legally(sheet_name)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                start_col = 2
                while start_col <= ws.max_column:
                    cell = ws.cell(row=row, column=start_col)
                    val = str(cell.value).strip() if cell.value is not None else ""

                    if val and val not in ["FREE", "BREAK"]:
                        raw_code = val.split(" ")[0]
                        code = raw_code.rstrip("T")
                        if code not in color_map:
                            color_map[code] = palette[color_index % len(palette)]
                            color_index += 1
                        fill = PatternFill(start_color=color_map[code], end_color=color_map[code], fill_type="solid")
                        cell.fill = fill

                        merge_count = 0
                        for col in range(start_col + 1, ws.max_column + 1):
                            next_cell = ws.cell(row=row, column=col)
                            if next_cell.value == cell.value:
                                next_cell.fill = fill
                                merge_count += 1
                            else:
                                break
                        if merge_count > 0:
                            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + merge_count)

                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        for col_idx in range(start_col, start_col + merge_count + 1):
                            ws.cell(row=row, column=col_idx).border = thin_border
                        start_col += merge_count + 1
                    elif val == "BREAK":
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        start_col += 1
                    else:
                        cell.border = thin_border
                        start_col += 1

            start_row = ws.max_row + 3
            headers = ["S.No", "Course Code", "Course Title", "L-T-P-S-C", "Faculty", "Color"]
            for idx, header in enumerate(headers, start=2):
                ws.cell(start_row, idx, header).border = thin_border
                ws.cell(start_row, idx).alignment = Alignment(horizontal="center", vertical="center")


            i = 1
            for code in color_map:
                if code.startswith("Elective_"):
                    continue
                ws.cell(start_row + i, 2, i).border = thin_border
                ws.cell(start_row + i, 3, code).border = thin_border
                course_name = next((c.title for c in self.courses if c.code == code), code)
                faculty = next((c.faculty for c in self.courses if c.code == code), "")
                ltpsc = next((c.ltp for c in self.courses if c.code == code), "")
                ws.cell(start_row + i, 4, course_name).border = thin_border
                ws.cell(start_row + i, 5, ltpsc).border = thin_border
                ws.cell(start_row + i, 5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(start_row + i, 6, faculty).border = thin_border
                ws.cell(start_row + i, 7, "").fill = PatternFill(
                    start_color=color_map[code],
                    end_color=color_map[code],
                    fill_type="solid"
                )
                ws.cell(start_row + i, 7).border = thin_border

                i += 1

            electives_header_row = start_row + i + 2
            e_headers = ["S.No", "Elective Basket", "Elective Title", "Faculty", "Room", "Color"]
            for idx, header in enumerate(e_headers, start=2):
                ws.cell(electives_header_row, idx, header).border = thin_border
                ws.cell(electives_header_row, idx).alignment = Alignment(horizontal="center", vertical="center")

            chosen_by_basket = {b: e for (b, e) in self.electives_by_sheet.get(sheet_name, [])}
            row_ctr = 1

            for basket in sorted(chosen_by_basket.keys()):
                elective_code = f"Elective_{basket}"
                chosen_e = chosen_by_basket[basket]

                all_electives = [c for c in self.courses if c.is_elective and c.basket == basket]

                for idx_e, e in enumerate(all_electives):
                    ws.cell(electives_header_row + row_ctr, 2, row_ctr).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 3, elective_code).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 4, e.title).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 5, e.faculty).border = thin_border
                    
                    key = f"{elective_code}||{e.title}"
                    room = self.elective_room_assignment.get(sheet_name, {}).get(key, "")
                    ws.cell(electives_header_row + row_ctr, 6, room).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 6).alignment = Alignment(horizontal="center", vertical="center")

                    ws.cell(electives_header_row + row_ctr, 7, "").fill = PatternFill(
                        start_color=color_map.get(elective_code, "FFFFFF"),
                        end_color=color_map.get(elective_code, "FFFFFF"),
                        fill_type="solid"
                    )
                    ws.cell(electives_header_row + row_ctr, 7).border = thin_border

                    row_ctr += 1

            timetable_max_row = len(self.days) + 1
            timetable_max_col = ws.max_column
            for row in ws.iter_rows(min_row=2, max_row=timetable_max_row, min_col=2, max_col=timetable_max_col):
                for cell in row:
                    cell.border = thin_border
            ws.freeze_panes = "B2"


            for col in ws.columns:
                max_length = 0
                try:
                    column_letter = col[0].column_letter
                except Exception:
                    continue
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"Formatted student timetable saved in {filename}")



    def _generate_faculty_workbook(self, faculty_filename):
        faculty_set = set()
        for c in self.courses:
            raw = c.faculty
            if raw:
                for p in [x.strip() for x in raw.split("/") if x.strip()]:
                    faculty_set.add(p)
        for ent in self.scheduled_entries:
            fstr = ent.get("faculty")
            if fstr:
                for p in [x.strip() for x in fstr.split("/") if x.strip()]:
                    faculty_set.add(p)

        faculty_tables = {}
        for f in faculty_set:
            faculty_tables[f] = {
                "First_Half": pd.DataFrame("    ", index=self.days, columns=self.slots),
                "Second_Half": pd.DataFrame("    ", index=self.days, columns=self.slots)
            }

        for ent in self.scheduled_entries:
            day = ent["day"]
            slot = ent["slot"]
            display = ent["display"]
            sheet = ent["sheet"]
            if ent.get("faculty"):
                faculties = [p.strip() for p in ent["faculty"].split("/") if p.strip()]
            else:
                matched = [c for c in self.courses if c.code == ent.get("code")]
                faculties = []
                for m in matched:
                    faculties.extend([p.strip() for p in m.faculty.split("/") if p.strip()])

            for f in faculties:
                faculty_tables[f][sheet].at[day, slot] = display

        with pd.ExcelWriter(faculty_filename, engine="openpyxl") as writer:
            for f in sorted(faculty_tables.keys()):
                safe = f[:31]
                df_first = faculty_tables[f]["First_Half"]
                df_second = faculty_tables[f]["Second_Half"]

                combined = pd.DataFrame()
                combined[" "] = [""] * (len(self.days) + 2)
                temp = df_first.reset_index()
                temp.columns = ["Day"] + list(df_first.columns)
                combined_first = temp

                spacer = pd.DataFrame([[""] * temp.shape[1]], columns=temp.columns)

                temp2 = df_second.reset_index()
                temp2.columns = ["Day"] + list(df_second.columns)
                combined_second = temp2

                final = pd.concat([combined_first, spacer, combined_second], ignore_index=True)
                final.to_excel(writer, sheet_name=safe, index=False)

        wb = load_workbook(faculty_filename)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        palette = ["FFC7CE","C6EFCE","FFEB9C","BDD7EE","D9EAD3","F4CCCC","D9D2E9","FCE5CD","C9DAF8","EAD1DC"]
        color_map = {}
        color_index = 0

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in range(2, ws.max_row + 1):
                start_col = 2
                while start_col <= ws.max_column:
                    cell = ws.cell(row=row, column=start_col)
                    if cell.value and str(cell.value).strip() not in ["FREE", ""]:
                        raw_code = str(cell.value).split(" ")[0].rstrip("T")
                        if raw_code not in color_map:
                            color_map[raw_code] = palette[color_index % len(palette)]
                            color_index += 1
                        fill = PatternFill(start_color=color_map[raw_code], end_color=color_map[raw_code], fill_type="solid")
                        cell.fill = fill
                        merge_count = 0
                        for col in range(start_col + 1, ws.max_column + 1):
                            next_cell = ws.cell(row=row, column=col)
                            if next_cell.value == cell.value:
                                next_cell.fill = fill
                                merge_count += 1
                            else:
                                break
                        if merge_count > 0:
                            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + merge_count)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        for col_idx in range(start_col, start_col + merge_count + 1):
                            ws.cell(row=row, column=col_idx).border = thin
                        start_col += merge_count + 1
                    else:
                        cell.border = thin
                        start_col += 1

                        
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin

            ws.freeze_panes = "B2"


            for col in ws.columns:
                max_length = 0
                try:
                    column = col[0].column_letter
                except:
                    continue
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2

        wb.save(faculty_filename)
        print(f"Saved faculty timetables to {faculty_filename}")

    def run_all_outputs(self, dept_name_prefix="CSE", student_filename=None, faculty_filename="faculty_timetable.xlsx"):
        
        if not student_filename:
            student_filename = f"{dept_name_prefix}_timetable.xlsx"

        
        self.scheduled_entries = []
        self.electives_by_sheet = {}
        self.elective_room_assignment = {}

        with pd.ExcelWriter(student_filename, engine="openpyxl") as writer:
           
            self.generate_timetable([c for c in self.courses if c.semester_half in ["1", "0"]], writer, "First_Half")
            self.generate_timetable([c for c in self.courses if c.semester_half in ["2", "0"]], writer, "Second_Half")

        if self.unscheduled_courses:
            unsched_file = f"{dept_name_prefix}_unscheduled_courses.xlsx"
            pd.DataFrame(self.unscheduled_courses).to_excel(unsched_file, index=False)
            print(f"Some courses couldn't be scheduled. See '{unsched_file}' for details.")
        wb = load_workbook(student_filename)
        for default in ["Sheet", "Sheet1"]:
            if default in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb[default])
        wb.save(student_filename)

        
        for sheet in wb.sheetnames:
            self._compute_elective_room_assignments_legally(sheet)

        
        self.format_student_timetable_with_legend(student_filename)

        
        self._generate_faculty_workbook(faculty_filename)


if __name__ == "__main__":
   
    departments = {
        "CSE-3-A": "data/CSE_3_A_courses.csv",
        "CSE-3-B": "data/CSE_3_B_courses.csv",
        "CSE-1-A": "data/CSE_1_A_courses.csv",
        "CSE-1-B": "data/CSE_1_B_courses.csv",
        "CSE-5-A": "data/CSE_5_A_courses.csv",
        "CSE-5-B": "data/CSE_5_B_courses.csv",
        "7-SEM": "data/DSAI_7_courses.csv",
        "DSAI-3": "data/DSAI_3_courses.csv",
        "ECE-3": "data/ECE_3_courses.csv",
        "DSAI-1": "data/DSAI_1_courses.csv",
        "ECE-1": "data/ECE_1_courses.csv",
        "DSAI-5": "data/DSAI_5_courses.csv",
        "ECE-5": "data/ECE_5_courses.csv",
    }
    rooms_file = "data/rooms.csv"
    slots_file = "data/timeslots.csv"
    global_room_usage = {}
    combined_faculty_filename = "faculty_timetable.xlsx"
    
    global_elective_slots = {}
    all_scheduled_entries = []

    for dept_name, course_file in departments.items():
        print(f"\nGenerating student timetable for {dept_name}...")
        scheduler = Scheduler(slots_file, course_file, rooms_file, global_room_usage, global_elective_slots)
        student_file = f"{dept_name}_timetable.xlsx"
        scheduler.run_all_outputs(dept_name_prefix=dept_name, student_filename=student_file, faculty_filename=combined_faculty_filename)

        all_scheduled_entries.extend(scheduler.scheduled_entries)
        for k, v in scheduler.course_room_map.items():    
            global_room_usage.setdefault("MAPPING", {})[k] = v    
    combined_courses = []
    for dept_name, course_file in departments.items():
        df = pd.read_csv(course_file)
        for _, row in df.iterrows():
            combined_courses.append(Course(row))
    helper = Scheduler(slots_file, departments[list(departments.keys())[0]], rooms_file, global_room_usage, global_elective_slots)
    helper.courses = combined_courses
    helper.scheduled_entries = all_scheduled_entries 
    helper._generate_faculty_workbook(combined_faculty_filename)
    print("\nAll done. Student timetables and combined faculty timetable generated.")