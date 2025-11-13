import re
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
ALLOWED_DATES = [
    "2025-11-20",
    "2025-11-21",
    "2025-11-22",
    "2025-11-23",
    "2025-11-24",
    "2025-11-25",
    "2025-11-26",
    "2025-11-27",
    "2025-11-28",
    "2025-11-29",
    "2025-11-30",
]

ALLOWED_DATES = [
    datetime.strptime(d, "%Y-%m-%d").date()
    for d in ALLOWED_DATES
]

SLOT_LABELS = ["09:00-12:00", "14:00-17:00"]
MAX_GLOBAL_EXAMS_PER_DAY = 4
MAX_EXAMS_PER_GROUP_PER_DAY = 1
DEFAULT_START_DATE = "2025-11-20"

ROOM_SORT_MODE = "small-first"  
USE_HALLS_LAST = True          

def invigilators_needed(capacity):
    return 2 if capacity >= 200 else 1

def extract_semester_id(group_name: str) -> str:
    m = re.search(r"(\d+)", str(group_name))
    return m.group(1) if m else str(group_name)

class Course:
    def __init__(self, row, group_name):
        self.group = group_name
        self.code = str(row["Course_Code"]).strip()
        self.title = str(row.get("Course_Title", self.code)).strip()
        try:
            self.students = int(str(row.get("Students", "0")).strip())
        except:
            self.students = 0
        flag = str(row.get("Elective", "0")).strip()
        self.is_elective = flag in ("1", "true", "True", "YES", "yes")
        self.basket = str(row.get("basket", "0")).strip()


class ExamScheduler:
    def __init__(self, rooms_file, departments, faculty_file, start_date=DEFAULT_START_DATE):
        self.rooms_df = pd.read_csv(rooms_file)
        self.departments = departments
        self.invig_df = pd.read_csv(faculty_file)
        self.start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        self.groups = list(departments.keys())
        self.invigilators = sorted([str(n).strip() for n in self.invig_df["Name"] if str(n).strip()])

        self.rooms = self._load_rooms()
        self.room_by_id = {r["Room_ID"]: r for r in self.rooms}

        self.courses = self._load_courses()

        self.room_remaining = {}
        self.group_daily = {}
        self.global_daily = {}
        self.used_rooms = {}

        self.scheduled = []
        self.unscheduled = []
        self.invig_assignments = []
        self._inv_idx = 0

    def _load_rooms(self):
        rooms = []
        for _, r in self.rooms_df.iterrows():
            rid = str(r["Room_ID"]).strip()
            t = str(r["Type"]).strip().lower()
            cap_raw = str(r["Capacity"]).strip()

            try:
                cap = int(cap_raw)
            except:
                cap = 0

            if cap <= 0:
                continue

            is_lab = ("lab" in t) or rid.upper().startswith(("L", "H"))
            is_library = ("library" in t)
            is_hall = ("hall" in t) or rid.upper() in {"C002", "C003", "C004"}

            if is_lab or is_library:
                continue

            usable = math.ceil(cap / 2)

            rooms.append({
                "Room_ID": rid,
                "Type": t,
                "Capacity": cap,
                "Usable": usable,
                "IsHall": bool(is_hall)
            })

        rooms.sort(key=lambda x: (x["Usable"], x["Room_ID"]))
        return rooms

    def _load_courses(self):
        out = {}
        for g, file in self.departments.items():
            df = pd.read_csv(file)
            lst = [Course(r, g) for _, r in df.iterrows() if int(str(r.get("Students", 0))) > 0]
            lst.sort(key=lambda c: (-c.students, c.code))
            out[g] = lst
        return out

    def _ensure_date(self, date):
        if date not in self.room_remaining:
            self.room_remaining[date] = {s: {r["Room_ID"]: r["Usable"] for r in self.rooms} for s in SLOT_LABELS}
        if date not in self.used_rooms:
            self.used_rooms[date] = {s: set() for s in SLOT_LABELS}
        if date not in self.group_daily:
            self.group_daily[date] = {g: 0 for g in self.groups}
        if date not in self.global_daily:
            self.global_daily[date] = 0

    def _ordered(self, ids, remaining):
        if ROOM_SORT_MODE == "small-first":
            return sorted(ids, key=lambda rid: (remaining.get(rid, 0), rid))
        else:
            return sorted(ids, key=lambda rid: (-remaining.get(rid, 0), rid))

    def _alloc_rooms(self, date, slot, need):
        remaining = self.room_remaining[date][slot]

        normal_ids = [r["Room_ID"] for r in self.rooms if not r["IsHall"]]
        hall_ids = [r["Room_ID"] for r in self.rooms if r["IsHall"]]

        def try_allocate(candidates):
            alloc = []
            total = 0
            for rid in self._ordered(candidates, remaining):
                avail = remaining.get(rid, 0)
                usable_cap = self.room_by_id[rid]["Usable"]
                avail = min(avail, usable_cap)
                if avail <= 0:
                    continue
                take = min(avail, need - total)
                if take > 0:
                    alloc.append((rid, take))
                    total += take
                if total >= need:
                    break
            return alloc if total >= need else None

        if USE_HALLS_LAST:
            alloc = try_allocate([rid for rid in normal_ids if remaining.get(rid, 0) > 0])
            if alloc is not None:
                return alloc
            alloc = try_allocate([rid for rid in normal_ids + hall_ids if remaining.get(rid, 0) > 0])
            return alloc
        else:
            all_ids = [rid for rid in normal_ids + hall_ids if remaining.get(rid, 0) > 0]
            return try_allocate(all_ids)

    def _book_alloc(self, date, slot, alloc):
        for rid, cnt in alloc:
            usable_cap = self.room_by_id[rid]["Usable"]
            safe_cnt = min(cnt, usable_cap)
            self.room_remaining[date][slot][rid] -= safe_cnt
            self.used_rooms[date][slot].add(rid)

    def _place_merged_course(self, code, title, students, groups_set, date, slot):
        if self.global_daily[date] >= MAX_GLOBAL_EXAMS_PER_DAY:
            return False
        for g in groups_set:
            if self.group_daily[date][g] >= MAX_EXAMS_PER_GROUP_PER_DAY:
                return False
        alloc = self._alloc_rooms(date, slot, students)
        if alloc is None:
            return False
        sanitized = []
        for rid, cnt in alloc:
            usable_cap = self.room_by_id[rid]["Usable"]
            sanitized.append((rid, min(cnt, usable_cap)))
        self._book_alloc(date, slot, sanitized)
        for g in groups_set:
            self.group_daily[date][g] += 1
        self.global_daily[date] += 1
        alloc_text = "; ".join([f"{rid}:{cnt}" for rid, cnt in sanitized])
        self.scheduled.append({
            "Date": date.strftime("%Y-%m-%d"),
            "Slot": slot,
            "Groups": ", ".join(sorted(groups_set)),
            "Course_Code": code,
            "Course_Title": title,
            "Students": students,
            "Allocations": alloc_text
        })
        return True

    def _plan_electives_by_semester(self):
        pool = {}
        for g in self.groups:
            sem = extract_semester_id(g)
            for c in self.courses[g]:
                if c.is_elective:
                    basket = getattr(c, "basket", "0")
                    if sem not in pool:
                        pool[sem] = {}
                    if basket not in pool[sem]:
                        pool[sem][basket] = {"electives": [], "groups": set()}
                    pool[sem][basket]["electives"].append(c)
                    pool[sem][basket]["groups"].add(g)
        return pool



    def _schedule_elective_block(self, sem, electives, groups_for_sem, start_day_offset, preferred_slot_index):
        day = start_day_offset
        while day < 300:
            date = self.start_date + timedelta(days=day)
            self._ensure_date(date)
            if self.global_daily[date] >= MAX_GLOBAL_EXAMS_PER_DAY:
                day += 1
                continue
            if any(self.group_daily[date][g] >= MAX_EXAMS_PER_GROUP_PER_DAY for g in groups_for_sem):
                day += 1
                continue
            slot = SLOT_LABELS[preferred_slot_index % len(SLOT_LABELS)]
            total_students = sum(c.students for c in electives)
            combined_alloc = self._alloc_rooms(date, slot, total_students)
            if combined_alloc is None:
                day += 1
                continue
            sanitized = []
            for rid, cnt in combined_alloc:
                usable_cap = self.room_by_id[rid]["Usable"]
                sanitized.append((rid, min(cnt, usable_cap)))
            self._book_alloc(date, slot, sanitized)
            self.global_daily[date] += 1
            for g in groups_for_sem:
                self.group_daily[date][g] += 1

            ordered_rooms = [(rid, cnt) for rid, cnt in sanitized]
            for c in electives:
                need = c.students
                per_elec_alloc = []
                for rid, seats in ordered_rooms:
                    if need <= 0:
                        break
                    take = min(need, seats)
                    if take > 0:
                        per_elec_alloc.append((rid, take))
                        need -= take
                alloc_text = "; ".join(f"{rid}:{cnt}" for rid, cnt in per_elec_alloc)
                self.scheduled.append({
                    "Date": date.strftime("%Y-%m-%d"),
                    "Slot": slot,
                    "Groups": c.group,
                    "Course_Code": c.code,
                    "Course_Title": c.title,
                    "Students": c.students,
                    "Allocations": alloc_text
                })
            return day + 1
        return day

    def _remove_scheduled_electives_from_pool(self):
        for g in self.groups:
            self.courses[g] = [c for c in self.courses[g] if not c.is_elective]

    def _all_done(self):
        return all(len(self.courses[g]) == 0 for g in self.groups)

    def generate(self):
        pool = self._plan_electives_by_semester()
        semesters = sorted(pool.keys(), key=lambda x: int(x))
        day_cursor = 0

        for sem in semesters:
            baskets = sorted(pool[sem].keys(), key=lambda x: int(x))
            for basket_idx, basket in enumerate(baskets):
                if day_cursor >= len(ALLOWED_DATES):
                    break
                date = ALLOWED_DATES[day_cursor]
                self._ensure_date(date)
                slot = SLOT_LABELS[basket_idx % len(SLOT_LABELS)]
                block = pool[sem][basket]
                groups = block["groups"]
                electives = block["electives"]
                total_students = sum(c.students for c in electives)
                if self.global_daily[date] >= MAX_GLOBAL_EXAMS_PER_DAY:
                    day_cursor += 1
                    continue
                if any(self.group_daily[date][g] >= MAX_EXAMS_PER_GROUP_PER_DAY for g in groups):
                    day_cursor += 1
                    continue
                alloc = self._alloc_rooms(date, slot, total_students)
                if alloc is None:
                    day_cursor += 1
                    continue
                self._book_alloc(date, slot, alloc)
                remaining_alloc = [(rid, cnt) for rid, cnt in alloc]
                for c in electives:
                    need = c.students
                    per_elec_alloc = []
                    for rid, seats in remaining_alloc:
                        if need <= 0:
                            break
                        take = min(seats, need)
                        if take > 0:
                            per_elec_alloc.append((rid, take))
                            need -= take
                    alloc_text = "; ".join(f"{rid}:{cnt}" for rid, cnt in per_elec_alloc)
                    self.scheduled.append({
                        "Date": date.strftime("%Y-%m-%d"),
                        "Slot": slot,
                        "Groups": ", ".join(sorted(groups)),
                        "Course_Code": c.code,
                        "Course_Title": c.title,
                        "Students": c.students,
                        "Allocations": alloc_text
                    })
                for g in groups:
                    self.group_daily[date][g] += 1
                self.global_daily[date] += 1

                if basket_idx % len(SLOT_LABELS) == len(SLOT_LABELS) - 1:
                    day_cursor += 1
            day_cursor += 1
        self._remove_scheduled_electives_from_pool()

        merged_regular = {}
        for g in self.groups:
            for c in self.courses[g]:
                if c.is_elective:
                    continue
                if c.code not in merged_regular:
                    merged_regular[c.code] = {
                        "code": c.code,
                        "title": c.title if c.title else c.code,
                        "students": 0,
                        "groups": set()
                    }
                merged_regular[c.code]["students"] += c.students
                merged_regular[c.code]["groups"].add(g)

        pending = sorted(merged_regular.values(), key=lambda x: (-x["students"], x["code"]))
        day = 0
        while pending and day < 300:
            if day >= len(ALLOWED_DATES):
                break

            date = ALLOWED_DATES[day]
            self._ensure_date(date)

            placed_today = 0
            si = 0
            i = 0
            while i < len(pending) and placed_today < MAX_GLOBAL_EXAMS_PER_DAY:
                exam = pending[i]
                if any(self.group_daily[date][g] >= MAX_EXAMS_PER_GROUP_PER_DAY for g in exam["groups"]):
                    i += 1
                    continue
                slot = SLOT_LABELS[si % len(SLOT_LABELS)]
                ok = self._place_merged_course(
                    code=exam["code"],
                    title=exam["title"],
                    students=exam["students"],
                    groups_set=exam["groups"],
                    date=date,
                    slot=slot
                )
                if ok:
                    pending.pop(i)
                    placed_today += 1
                    si += 1
                else:
                    i += 1
            day += 1

        if pending:
            for exam in pending:
                self.unscheduled.append({
                    "Group": ", ".join(sorted(exam["groups"])),
                    "Course_Code": exam["code"],
                    "Course_Title": exam["title"],
                    "Students": exam["students"]
                })

        self._assign_invigilators()

    def _assign_invigilators(self):
        for d in sorted(self.used_rooms.keys()):
            assigned_today = set()
            for slot in SLOT_LABELS:
                rooms = sorted(list(self.used_rooms[d][slot]))
                for rid in rooms:
                    cap = self.room_by_id[rid]["Capacity"]
                    k = invigilators_needed(cap)
                    picks = []
                    while len(picks) < k and self.invigilators:
                        name = self.invigilators[self._inv_idx % len(self.invigilators)]
                        self._inv_idx += 1
                        if name not in assigned_today:
                            picks.append(name)
                            assigned_today.add(name)
                    exam_names = []
                    date_str = d.strftime("%Y-%m-%d")
                    for rec in self.scheduled:
                        if rec["Date"] == date_str and rec["Slot"] == slot:
                            for a in rec["Allocations"].split(";"):
                                if rid == a.split(":")[0]:
                                    exam_names.append(f"{rec['Course_Code']}")
                                    break
                    self.invig_assignments.append({
                        "Date": date_str,
                        "Slot": slot,
                        "Room_ID": rid,
                        "Exam": " | ".join(sorted(set(exam_names))),
                        "Invigilators": ", ".join(picks)
                    })

    def _fmt(self, file):
        wb = load_workbook(file)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for s in wb.sheetnames:
            ws = wb[s]
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin
                    if cell.row == 1 or cell.column == 1:
                        cell.fill = gray
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col in ws.columns:
                mx = 12
                for cell in col:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(mx + 4, 70)
            ws.row_dimensions[1].height = 24
        wb.save(file)

    def _parse_alloc(self, s):
        out = {}
        if not s:
            return out
        parts = [p.strip() for p in str(s).split(";") if p.strip()]
        for p in parts:
            if ":" in p:
                rid, cnt = p.split(":")
                try:
                    out[rid.strip()] = out.get(rid.strip(), 0) + int(cnt.strip())
                except:
                    pass
        return out

    def _format_alloc(self, alloc_dict):
        order = [r["Room_ID"] for r in self.rooms]
        items = [(rid, alloc_dict[rid]) for rid in order if rid in alloc_dict and alloc_dict[rid] > 0]
        if not items:
            items = sorted(alloc_dict.items(), key=lambda x: x[0])
        return "; ".join(f"{rid}:{cnt}" for rid, cnt in items)

    def _build_merged(self):
        rows = self.scheduled
        groups = {}
        title_map = {}

        for r in rows:
            k = (r["Date"], r["Slot"], r["Course_Code"])
            title_map[r["Course_Code"]] = r.get("Course_Title", r["Course_Code"])

            if k not in groups:
                groups[k] = {"Students": 0, "Alloc": {}, "Groups": set()}

            groups[k]["Students"] += int(r.get("Students", 0) or 0)

            alloc_dict = self._parse_alloc(r.get("Allocations", ""))

            for rid, cnt in alloc_dict.items():
                if rid not in groups[k]["Alloc"]:
                    groups[k]["Alloc"][rid] = 0

                if len(groups[k]["Groups"]) > 1:
                    groups[k]["Alloc"][rid] = max(groups[k]["Alloc"].get(rid, 0), cnt)
                else:
                    groups[k]["Alloc"][rid] = groups[k]["Alloc"].get(rid, 0) + cnt

            gs = str(r.get("Groups", "")).strip()
            if gs:
                for gname in [x.strip() for x in gs.split(",") if x.strip()]:
                    groups[k]["Groups"].add(gname)

        for (date, slot, code), v in groups.items():
            for rid in list(v["Alloc"].keys()):
                if rid in self.room_by_id:
                    v["Alloc"][rid] = min(v["Alloc"][rid], self.room_by_id[rid]["Usable"])

        merged_rows = []
        for (date, slot, code), v in sorted(groups.items()):
            merged_rows.append({
                "Date": date,
                "Slot": slot,
                "Course_Code": code,
                "Students": v["Students"],
                "Allocations": self._format_alloc(v["Alloc"]),
                "Groups": ", ".join(sorted(v["Groups"])) if v["Groups"] else ""
            })

        legend = sorted([(code, title) for code, title in title_map.items()], key=lambda x: x[0])

        return pd.DataFrame(merged_rows), pd.DataFrame(legend, columns=["Course_Code", "Course_Title"])


    def _build_grid(self, merged_df):
        dates = sorted(merged_df["Date"].unique())
        grid = pd.DataFrame(index=SLOT_LABELS, columns=dates)
        for d in dates:
            for s in SLOT_LABELS:
                subset = merged_df[(merged_df["Date"] == d) & (merged_df["Slot"] == s)]
                if subset.empty:
                    grid.at[s, d] = ""
                else:
                    codes = subset["Course_Code"].tolist()
                    grid.at[s, d] = ", ".join(codes)
        return grid

    def export(self, out="exam_timetables.xlsx", uns="unscheduled_exams.xlsx", invig="invigilation.xlsx"):
        merged_df, legend_df = self._build_merged()
        grid_df = self._build_grid(merged_df)
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            merged_df.to_excel(w, sheet_name="Merged", index=False)
            grid_df.to_excel(w, sheet_name="Grid", index=True)
            legend_df.to_excel(w, sheet_name="Legend", index=False)
        self._fmt(out)
        if self.unscheduled:
            pd.DataFrame(self.unscheduled).to_excel(uns, index=False)
        if self.invig_assignments:
            pd.DataFrame(self.invig_assignments).sort_values(by=["Date", "Slot", "Room_ID"]).to_excel(invig, index=False)

def run_example():
    departments = {
        "CSE-3": "data/exam_data/CSE_3.csv",
        "ECE-3": "data/exam_data/ECE_3.csv",
        "DSAI-3": "data/exam_data/DSAI_3.csv",
        "CSE-1": "data/exam_data/CSE_1.csv",
        "ECE-1": "data/exam_data/ECE_1.csv",
        "DSAI-1": "data/exam_data/DSAI_1.csv",
        "CSE-5": "data/exam_data/CSE_5.csv",
        "ECE-5": "data/exam_data/ECE_5.csv", 
        "DSAI-5": "data/exam_data/DSAI_5.csv",
        "DSAI-7": "data/exam_data/DSAI_7.csv",

    }
    rooms = "data/exam_data/rooms.csv"
    faculty = "data/exam_data/faculty.csv"
    s = ExamScheduler(rooms, departments, faculty)
    s.generate()
    s.export()

if __name__ == "__main__":
    run_example()
