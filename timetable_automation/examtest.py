import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta

SLOT_LABELS = ["09:00-12:00", "14:00-17:00"]
MAX_GLOBAL_EXAMS_PER_DAY = 3
MAX_EXAMS_PER_GROUP_PER_DAY = 1
DEFAULT_START_DATE = "2025-12-01"

class Course:
    def __init__(self, row, group_name):
        self.group = group_name
        self.code = str(row["Course_Code"]).strip()
        self.title = str(row.get("Course_Title", self.code)).strip()
        self.faculty = str(row.get("Faculty", "")).strip()
        self.semester_half = str(row.get("Semester_Half", "0")).strip()
        self.is_elective = str(row.get("Elective", 0)).strip() == "1"
        try:
            self.students = int(str(row.get("Students", "0")).strip())
        except:
            self.students = 0

class ExamScheduler:
    def __init__(self, rooms_file, departments, start_date=DEFAULT_START_DATE):
        self.rooms_df = pd.read_csv(rooms_file)
        self.departments = departments
        self.start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        self.groups = list(departments.keys())
        self.rooms = self._load_rooms()
        self.courses = self._load_courses()
        self.room_usage = {}
        self.group_daily = {}
        self.global_daily = {}
        self.scheduled = []
        self.unscheduled = []

    def _load_rooms(self):
        rooms = []
        for _, r in self.rooms_df.iterrows():
            rid = str(r["Room_ID"]).strip()
            t = str(r["Type"]).strip().lower()
            try:
                cap = int(str(r["Capacity"]).strip())
            except:
                cap = None
            if not cap or cap <= 0:
                continue
            if "lab" in t or "library" in t:
                continue
            rooms.append({"Room_ID": rid, "Capacity": cap})
        rooms.sort(key=lambda x: (-x["Capacity"], x["Room_ID"]))
        return rooms

    def _load_courses(self):
        out = {}
        for g, file in self.departments.items():
            df = pd.read_csv(file)
            lst = [Course(r, g) for _, r in df.iterrows()]
            lst = [c for c in lst if c.students > 0]
            lst.sort(key=lambda c: (-c.students, c.code))
            out[g] = lst
        return out

    def _ensure_date(self, date):
        if date not in self.room_usage:
            self.room_usage[date] = {s: set() for s in SLOT_LABELS}
        if date not in self.group_daily:
            self.group_daily[date] = {g: 0 for g in self.groups}
        if date not in self.global_daily:
            self.global_daily[date] = 0

    def _alloc_rooms(self, date, slot, need):
        used = self.room_usage[date][slot]
        chosen = []
        total = 0
        for r in self.rooms:
            if r["Room_ID"] in used:
                continue
            chosen.append(r["Room_ID"])
            total += r["Capacity"]
            if total >= need:
                break
        if total < need:
            return None
        return chosen

    def _place(self, course, date, slot):
        if self.global_daily[date] >= MAX_GLOBAL_EXAMS_PER_DAY:
            return False
        if self.group_daily[date][course.group] >= MAX_EXAMS_PER_GROUP_PER_DAY:
            return False
        rooms = self._alloc_rooms(date, slot, course.students)
        if rooms is None:
            return False
        for r in rooms:
            self.room_usage[date][slot].add(r)
        self.group_daily[date][course.group] += 1
        self.global_daily[date] += 1
        self.scheduled.append({
            "Date": date.strftime("%Y-%m-%d"),
            "Slot": slot,
            "Group": course.group,
            "Course_Code": course.code,
            "Course_Title": course.title,
            "Students": course.students,
            "Rooms": ", ".join(rooms)
        })
        return True

    def _all_done(self):
        for g in self.groups:
            if self.courses[g]:
                return False
        return True

    def generate(self):
        day = 0
        while True:
            if self._all_done():
                break
            date = self.start_date + timedelta(days=day)
            self._ensure_date(date)
            pending_groups = [g for g in self.groups if self.courses[g]]
            pending_groups = pending_groups[:MAX_GLOBAL_EXAMS_PER_DAY]
            slots_cycle = SLOT_LABELS * 2
            si = 0
            for g in pending_groups:
                c = self.courses[g][0]
                slot = slots_cycle[si]
                ok = self._place(c, date, slot)
                if ok:
                    self.courses[g].pop(0)
                    si += 1
            day += 1
            if day > 120:
                break
        for g in self.groups:
            for c in self.courses[g]:
                self.unscheduled.append({
                    "Group": g,
                    "Course_Code": c.code,
                    "Course_Title": c.title,
                    "Students": c.students
                })

    def _df_group(self, recs):
        dates = sorted({r["Date"] for r in recs})
        df = pd.DataFrame("", index=dates, columns=SLOT_LABELS)
        for r in recs:
            df.at[r["Date"], r["Slot"]] = f"{r['Course_Code']} - {r['Course_Title']} [{r['Students']}] @ {r['Rooms']}"
        df.index.name = "Date"
        return df

    def _fmt(self, file):
        wb = load_workbook(file)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for s in wb.sheetnames:
            ws = wb[s]
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin
                    if cell.row == 1 or cell.column == 1:
                        cell.fill = gray
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col in ws.columns:
                l = col[0].column_letter
                mx = 12
                for cell in col:
                    if cell.value:
                        mx = max(mx, len(str(cell.value)))
                ws.column_dimensions[l].width = min(mx + 4, 60)
            ws.row_dimensions[1].height = 24
        wb.save(file)

    def export(self, out="exam_timetables.xlsx", uns="unscheduled_exams.xlsx"):
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            dfm = pd.DataFrame(self.scheduled)
            if not dfm.empty:
                dfm.sort_values(by=["Date","Slot","Group","Course_Code"]).to_excel(w, sheet_name="Master", index=False)
            else:
                pd.DataFrame(columns=["Date","Slot","Group","Course_Code","Course_Title","Students","Rooms"]).to_excel(w, sheet_name="Master", index=False)
            for g in self.groups:
                rec = [r for r in self.scheduled if r["Group"] == g]
                df = self._df_group(rec)
                df.to_excel(w, sheet_name=g[:31], index=True)
        self._fmt(out)
        if self.unscheduled:
            pd.DataFrame(self.unscheduled).to_excel(uns, index=False)

def run_example():
    departments = {
        "CSE-3-A": "data/CSE_3_courses-A.csv",
        "CSE-3-B": "data/CSE_3_courses-B.csv",
        "DSAI-3": "data/DSAI_courses.csv"
    }
    rooms = "data/rooms.csv"
    s = ExamScheduler(rooms, departments)
    s.generate()
    s.export()

if __name__ == "__main__":
    run_example()
