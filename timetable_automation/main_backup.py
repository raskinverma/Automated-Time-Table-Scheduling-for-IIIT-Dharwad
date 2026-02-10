import pandas as pd
import random
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill
RANDOM_SEED = 42
random.seed(RANDOM_SEED)

class Course:
    def __init__(self, row):
        self.code = str(row["Course_Code"]).strip()
        self.basket = int(row.get("basket", 0))
        self.title = str(row.get("Course_Title", self.code)).strip()
        self.faculty = str(row.get("Faculty", "")).strip()
        self.ltp = str(row["L-T-P-S-C"]).strip()
        self.semester_half = str(row.get("Semester_Half", "0")).strip()
        elective_raw = row.get("Elective", 0)
        elective_str = str(elective_raw).strip().lower()
        elective_num = 0.0
        try:
            elective_num = float(elective_str) if elective_str else 0.0
        except Exception:
            elective_num = 1.0 if elective_str in {"true", "yes", "y"} else 0.0
        self.is_elective = (elective_num > 0) or (self.basket > 0)
        raw_combined = row.get("is_combined", row.get("Is_Combined", 0))
        self.is_combined = str(raw_combined).strip().lower() in {"1", "true", "yes", "y"}
        students_raw = row.get("Students", row.get("students", 0))
        try:
            self.students = max(0, int(float(str(students_raw).strip())))
        except Exception:
            self.students = 0
        try:
            self.L, self.T, self.P, self.S, self.C = map(int, self.ltp.split("-"))
        except Exception:
            self.L, self.T, self.P = 0, 0, 0


class Scheduler:
    def __init__(
        self,
        slots_file,
        courses_file,
        rooms_file,
        global_room_usage,
        global_elective_slots=None,
        dept_name="",
        global_elective_slot_usage=None,
        global_elective_room_templates=None,
        global_elective_room_usage=None,
        global_elective_representatives=None,
        global_combined_slots=None,
        global_combined_room_usage=None,
        global_combined_strength=None,
        global_c004_reserved_slots=None,
    ):
        df = pd.read_csv(slots_file)
        self.slots = [f"{row['Start_Time'].strip()}-{row['End_Time'].strip()}" for _, row in df.iterrows()]
        self.slot_durations = {s: self._slot_duration(s) for s in self.slots}

        courses_df = pd.read_csv(courses_file)
        self.courses = [Course(row) for _, row in courses_df.iterrows()]

        rooms_df = pd.read_csv(rooms_file)
        self.classrooms = []
        self.labs = []
        self.all_rooms = []
        self.room_capacity = {}
        for _, row in rooms_df.iterrows():
            room_id = str(row["Room_ID"]).strip()
            self.all_rooms.append(room_id)
            try:
                cap = max(0, int(float(str(row.get("Capacity", 0)).strip())))
            except Exception:
                cap = 0
            self.room_capacity[room_id.upper()] = cap
            if room_id.upper().startswith("L"):
                self.labs.append(room_id)
            elif room_id.upper().startswith("C"):
                self.classrooms.append(room_id)
            else:
                continue
        # Room policy:
        # - C002/C003 are reserved for compulsory courses.
        # - C004 is reserved for combined courses only.
        self.compulsory_only_classrooms = {"C002", "C003"}
        self.non_compulsory_blocked_classrooms = {"C004"}

        self.days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        self.excluded_slots = ["07:30-09:00", "13:15-14:00"]
        self.MAX_ATTEMPTS = 2000
        self.unscheduled_courses = []
        self.course_room_map = {}
        self.global_room_usage = global_room_usage
        self.scheduled_entries = []
        self.electives_by_sheet = {}
        self.elective_room_assignment = {}
        self.break_length_slots = 1
        self.global_elective_slots = global_elective_slots if global_elective_slots is not None else {}
        self.dept_name = str(dept_name).strip()
        match = re.search(r"\d+", self.dept_name)
        self.semester_group = match.group(0) if match else "UNKNOWN"
        self.global_elective_slot_usage = global_elective_slot_usage if global_elective_slot_usage is not None else {}
        self.global_elective_room_templates = (
            global_elective_room_templates if global_elective_room_templates is not None else {}
        )
        self.global_elective_room_usage = (
            global_elective_room_usage if global_elective_room_usage is not None else {}
        )
        self.global_elective_representatives = (
            global_elective_representatives if global_elective_representatives is not None else {}
        )
        self.global_combined_slots = global_combined_slots if global_combined_slots is not None else {}
        self.global_combined_room_usage = (
            global_combined_room_usage if global_combined_room_usage is not None else {}
        )
        self.global_combined_strength = (
            global_combined_strength if global_combined_strength is not None else {}
        )
        self.global_c004_reserved_slots = (
            global_c004_reserved_slots if global_c004_reserved_slots is not None else {}
        )
        self.dept_prefix = self.dept_name.split("-")[0].strip().upper() if self.dept_name else ""
        self.combined_cluster_id = self._resolve_combined_cluster()
        # Soft constraint: prefer avoiding cross-sem elective overlap, but relax if needed.
        self.relax_cross_sem_elective_block = True
        self._bootstrap_c004_reserved_slots_from_templates()

    def _elective_template_key(self, basket_id, session_type, sheet_name):
        # Share elective templates across all branches of the same semester.
        return (
            self.semester_group,
            sheet_name,
            basket_id,
            session_type,
        )

    def _elective_representative_key(self, basket_id, sheet_name):
        return (
            self.semester_group,
            sheet_name,
            basket_id,
        )

    def _pick_elective_representative(self, basket_id, group, sheet_name):
        if not group:
            return None
        ordered = sorted(group, key=lambda c: (c.code, c.title, c.faculty))
        key = self._elective_representative_key(basket_id, sheet_name)
        saved = self.global_elective_representatives.get(key)

        if isinstance(saved, dict):
            code = str(saved.get("code", "")).strip().upper()
            title = str(saved.get("title", "")).strip().lower()
            ltp = str(saved.get("ltp", "")).strip()
            if code:
                for candidate in ordered:
                    if str(candidate.code).strip().upper() == code:
                        return candidate
            if title:
                for candidate in ordered:
                    if str(candidate.title).strip().lower() == title and (not ltp or candidate.ltp == ltp):
                        return candidate
            if ltp:
                for candidate in ordered:
                    if candidate.ltp == ltp:
                        return candidate

        chosen = ordered[0]
        self.global_elective_representatives.setdefault(
            key,
            {
                "code": chosen.code,
                "title": chosen.title,
                "ltp": chosen.ltp,
            },
        )
        return chosen

    def _course_in_sheet_half(self, course, sheet_name):
        half = str(getattr(course, "semester_half", "0")).strip()
        if sheet_name == "First_Half":
            return half in {"1", "0"}
        if sheet_name == "Second_Half":
            return half in {"2", "0"}
        return True

    def _resolve_combined_cluster(self):
        cluster_groups = (
            {"CSE"},
            {"DSAI", "ECE"},
        )
        for group in cluster_groups:
            if self.dept_prefix in group:
                return "+".join(sorted(group))
        return self.dept_prefix if self.dept_prefix else self.dept_name.upper()

    def _combined_template_key(self, code, session_type, sheet_name):
        return (self.semester_group, self.combined_cluster_id, sheet_name, code, session_type)

    def _record_combined_slots(self, template_key, day, slots, room):
        entries = self.global_combined_slots.setdefault(template_key, [])
        signature = (day, tuple(slots), room)
        for ent in entries:
            if (ent.get("day"), tuple(ent.get("slots", [])), ent.get("room", "")) == signature:
                return
        entries.append({"day": day, "slots": list(slots), "room": room})

    def _room_matches_session(self, room_id, session_type):
        room_upper = room_id.upper()
        if session_type == "P":
            return room_upper.startswith("L")
        return not room_upper.startswith("L")

    def _room_allowed_for_course(self, room_id, is_compulsory, is_combined_course=False):
        room_upper = str(room_id).strip().upper()
        # Hard rule: C004 can only be used by combined courses.
        if room_upper == "C004" and not is_combined_course:
            return False
        if room_upper in self.compulsory_only_classrooms and not is_compulsory:
            return False
        if room_upper in self.non_compulsory_blocked_classrooms and not is_compulsory:
            return False
        return True

    def _room_has_capacity(self, room_id, min_capacity_needed):
        if not min_capacity_needed or min_capacity_needed <= 0:
            return True
        return self.room_capacity.get(str(room_id).strip().upper(), 0) >= min_capacity_needed

    def _combined_strength_key(self, code):
        return (self.semester_group, self.combined_cluster_id, str(code).strip().upper())

    def _required_capacity_for_course(self, course, is_elective, is_combined):
        if is_elective or not is_combined:
            return None
        own_strength = max(0, int(getattr(course, "students", 0) or 0))
        combined_strength = max(
            0,
            int(self.global_combined_strength.get(self._combined_strength_key(course.code), 0) or 0),
        )
        need = max(own_strength, combined_strength)
        return need if need > 0 else None

    def _bootstrap_c004_reserved_slots_from_templates(self):
        sem_slots = self.global_c004_reserved_slots.setdefault(self.semester_group, set())
        for key, entries in self.global_combined_slots.items():
            if not isinstance(key, tuple) or not key:
                continue
            if str(key[0]) != str(self.semester_group):
                continue
            for ent in entries:
                room = str(ent.get("room", "")).strip().upper()
                if room != "C004":
                    continue
                day = ent.get("day")
                for slot in ent.get("slots", []):
                    sem_slots.add((day, slot))

    def _reserve_c004_slots(self, day, slots):
        sem_slots = self.global_c004_reserved_slots.setdefault(self.semester_group, set())
        for slot in slots:
            sem_slots.add((day, slot))

    def _is_c004_available_for_course_slots(self, day, slots, room_id, is_combined_course):
        if str(room_id).strip().upper() != "C004":
            return True
        return bool(is_combined_course)

    def _is_room_available(self, day, slots, room_id, combined_key=None, sheet_name=None):
        room_usage = self._sheet_scoped_usage(self.global_room_usage, sheet_name)
        combined_room_usage = self._sheet_scoped_usage(self.global_combined_room_usage, sheet_name)
        for slot in slots:
            used_rooms = room_usage.get(day, {}).get(slot, [])
            if room_id not in used_rooms:
                continue
            if not combined_key:
                return False
            owner = combined_room_usage.get(day, {}).get(slot, {}).get(room_id)
            if owner != combined_key:
                return False
        return True

    def _combined_day_order(self):
        sem = str(self.semester_group).strip()
        if sem == "1":
            return ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        if sem == "3":
            return ["Thursday", "Friday", "Wednesday", "Tuesday", "Monday"]
        if sem == "5":
            return ["Wednesday", "Thursday", "Friday", "Monday", "Tuesday"]
        return self.days.copy()

    def _sheet_scoped_usage(self, usage_map, sheet_name):
        # First/Second half run in separate windows; keep their room usage independent.
        if sheet_name in {"First_Half", "Second_Half"}:
            return usage_map.setdefault(sheet_name, {})
        return usage_map

    def _pick_room_for_slots(
        self,
        day,
        slots,
        code,
        session_type,
        sheet_name=None,
        combined_key=None,
        preferred_room=None,
        is_compulsory=True,
        min_capacity_needed=None,
        is_combined_course=False,
    ):
        def room_ok(room_id):
            return (
                self._room_matches_session(room_id, session_type)
                and self._room_allowed_for_course(
                    room_id,
                    is_compulsory,
                    is_combined_course=is_combined_course,
                )
                and self._room_has_capacity(room_id, min_capacity_needed)
                and self._is_c004_available_for_course_slots(
                    day, slots, room_id, is_combined_course
                )
                and self._is_room_available(
                    day,
                    slots,
                    room_id,
                    combined_key=combined_key,
                    sheet_name=sheet_name,
                )
            )

        if preferred_room and room_ok(preferred_room):
            return preferred_room

        # For combined lecture/tutorial sessions, prefer C004 first.
        if is_combined_course and session_type != "P":
            c004_room = next(
                (r for r in self.classrooms if str(r).strip().upper() == "C004"),
                "",
            )
            if c004_room and room_ok(c004_room):
                return c004_room

        mapped = self.course_room_map.get(code)
        if mapped and room_ok(mapped):
            return mapped

        possible_rooms = self.labs if session_type == "P" else self.classrooms
        room_candidates = possible_rooms.copy()
        if is_combined_course and session_type != "P":
            c004_exact = [r for r in room_candidates if str(r).strip().upper() == "C004"]
            others = [r for r in room_candidates if str(r).strip().upper() != "C004"]
            random.shuffle(others)
            room_candidates = c004_exact + others
        else:
            random.shuffle(room_candidates)
        for room_id in room_candidates:
            if room_ok(room_id):
                return room_id
        return ""

    def _is_blocked_elective_slot(self, day, slot):
        for sem, used in self.global_elective_slot_usage.items():
            if sem != self.semester_group and (day, slot) in used:
                return True
        return False

    def _reserve_elective_slots(self, day, slots):
        sem_slots = self.global_elective_slot_usage.setdefault(self.semester_group, set())
        for s in slots:
            sem_slots.add((day, s))


 
    def _slot_duration(self, slot):
        start, end = slot.split("-")
        h1, m1 = map(int, start.split(":"))
        h2, m2 = map(int, end.split(":"))
        return (h2 + m2 / 60) - (h1 + m1 / 60)

    def _get_free_blocks(self, timetable, day):
        free_blocks, block = [], []
        for slot in self.slots:
            if timetable.at[day, slot] == "" and slot not in self.excluded_slots:
                block.append(slot)
            else:
                if block:
                    free_blocks.append(block)
                    block = []
        if block:
            free_blocks.append(block)
        return free_blocks

    def _allocate_session(
        self,
        timetable,
        lecturer_busy,
        labs_scheduled,
        day,
        faculty,
        code,
        duration_hours,
        session_type="L",
        is_elective=False,
        sheet_name=None,
        force_slots=None,
        min_rooms_needed=1,
        relax_elective_block=False,
        combined_key=None,
        preferred_room=None,
        min_capacity_needed=None,
    ):
       
        for entry in self.scheduled_entries:
            if entry["day"] == day and entry["code"] == code and entry["sheet"] == sheet_name:
                return None
       
        
        if session_type == "P" and labs_scheduled[day]:
            return None

        # Try to find valid slots
        valid_slots_found = None
        room_to_use = ""
        room_usage = self._sheet_scoped_usage(self.global_room_usage, sheet_name)
        combined_room_usage = self._sheet_scoped_usage(self.global_combined_room_usage, sheet_name)

        if force_slots:
            # FORCE MODE: Check if provided slots are free in local timetable
            conflict = False
            for s in force_slots:
                if timetable.at[day, s] != "":
                    conflict = True
                    break
                if is_elective and not relax_elective_block and self._is_blocked_elective_slot(day, s):
                    conflict = True
                    break
            
            if not conflict:
                # Check faculty busy
                if faculty:
                    day_busy = lecturer_busy.get(day, {})
                    if isinstance(day_busy, dict):
                        if any(faculty in day_busy.get(s, []) for s in force_slots):
                            conflict = True
                    elif faculty in day_busy:
                        conflict = True
                
            if not conflict:
                valid_slots_found = force_slots
                if is_elective and min_rooms_needed > 1:
                     # Verify capacity even for forced slots (though typically they are pre-validated)
                     if any(len(self.all_rooms) - len(room_usage.get(day, {}).get(s, [])) < min_rooms_needed for s in force_slots):
                         valid_slots_found = None

        else:
            # SEARCH MODE: Find all candidate sub-blocks that fit the duration
            free_blocks = self._get_free_blocks(timetable, day)
            candidates = []

            for block in free_blocks:
                # Sliding window over the block
                for i in range(len(block)):
                    dur_accum = 0
                    current_slots = []
                    for j in range(i, len(block)):
                        s = block[j]
                        current_slots.append(s)
                        dur_accum += self.slot_durations[s]
                        
                        if dur_accum >= duration_hours:
                            # Found a valid sub-block
                            waste = dur_accum - duration_hours
                            candidates.append({
                                'slots': current_slots,
                                'waste': waste,
                                'start_idx': i # mainly for stability if needed
                            })
                            # Once we reach required duration, we can stop extending strictly for "minimum fit"
                            # But if the next slot creates a "better" fit (unlikely if duration increases), we'd continue.
                            # Since slot durations are positive, adding more slots only increases waste. 
                            # So we stop this inner extension loop.
                            break 
            
            # Sort candidates by waste (ascending) to prefer tight fits
            candidates.sort(key=lambda x: x['waste'])

            for cand in candidates:
                current_slots = cand['slots']
                
                # Check faculty availability
                if faculty:
                    day_busy = lecturer_busy.get(day, {})
                    if isinstance(day_busy, dict):
                        if any(faculty in day_busy.get(s, []) for s in current_slots):
                            continue
                    elif faculty in day_busy:
                        continue

                # Check room capacity for electives
                if is_elective and min_rooms_needed > 1:
                     if any(len(self.all_rooms) - len(room_usage.get(day, {}).get(s, [])) < min_rooms_needed for s in current_slots):
                         continue
                if is_elective and not relax_elective_block and any(self._is_blocked_elective_slot(day, s) for s in current_slots):
                    continue

                # Check Room Availability
                chosen_room = ""
                if not is_elective:
                    chosen_room = self._pick_room_for_slots(
                        day,
                        current_slots,
                        code,
                        session_type,
                        sheet_name=sheet_name,
                        combined_key=combined_key,
                        preferred_room=preferred_room,
                        is_compulsory=not is_elective,
                        min_capacity_needed=min_capacity_needed,
                        is_combined_course=bool(combined_key),
                    )
                    if not chosen_room:
                        continue

                # Found a valid candidate!
                valid_slots_found = current_slots
                room_to_use = chosen_room
                break
        
        if not valid_slots_found:
            return None

        if not is_elective and not room_to_use:
            room_to_use = self._pick_room_for_slots(
                day,
                valid_slots_found,
                code,
                session_type,
                sheet_name=sheet_name,
                combined_key=combined_key,
                preferred_room=preferred_room,
                is_compulsory=not is_elective,
                min_capacity_needed=min_capacity_needed,
                is_combined_course=bool(combined_key),
            )
            if not room_to_use:
                return None
        if (
            not is_elective
            and str(room_to_use).strip().upper() == "C004"
            and not combined_key
        ):
            return None

        # Apply allocation
        slots_to_use = valid_slots_found
        
        if not is_elective:
            if not self.course_room_map.get(code) and room_to_use:
                 self.course_room_map[code] = room_to_use
            
            if room_to_use:
                for s in slots_to_use:
                     day_slots = room_usage.setdefault(day, {}).setdefault(s, [])
                     if room_to_use not in day_slots:
                         day_slots.append(room_to_use)
                     if combined_key:
                         combined_room_usage.setdefault(day, {}).setdefault(s, {})[room_to_use] = combined_key
                if combined_key and str(room_to_use).strip().upper() == "C004":
                    self._reserve_c004_slots(day, slots_to_use)

        for i, s in enumerate(slots_to_use):
            if session_type == "L":
                display_text = f"{code} ({room_to_use})" if (room_to_use and not is_elective) else code
            elif session_type == "T":
                display_text = f"{code}T ({room_to_use})" if (room_to_use and not is_elective) else f"{code}T"
            elif session_type == "P":
                suffix = f" (Lab-{room_to_use})" if room_to_use and not is_elective else " (Lab)"
                display_text = f"{code}{suffix}"
            else:
                display_text = code

            timetable.at[day, s] = display_text

            self.scheduled_entries.append(
                {
                    "sheet": sheet_name,
                    "day": day,
                    "slot": s,
                    "code": code,
                    "display": display_text,
                    "faculty": faculty,
                    "room": room_to_use,
                }
            )

            # Mark gap slot if applies
            if i < len(slots_to_use) - 1:
                idx = self.slots.index(s)
                if idx + 1 < len(self.slots):
                    gap_slot = self.slots[idx + 1]
                    if timetable.at[day, gap_slot] == "" and self.slot_durations[gap_slot] == 0.25:
                        timetable.at[day, gap_slot] = "FREE"

        if faculty:
            day_busy = lecturer_busy.get(day, {})
            if isinstance(day_busy, dict):
                day_busy = lecturer_busy.setdefault(day, {})
                for s in slots_to_use:
                    day_busy.setdefault(s, []).append(faculty)
            else:
                day_busy = lecturer_busy.setdefault(day, day_busy if isinstance(day_busy, list) else [])
                if faculty not in day_busy:
                    day_busy.append(faculty)

        if is_elective:
            self._reserve_elective_slots(day, slots_to_use)

        if session_type == "P":
            labs_scheduled[day] = True
        
        # Add break
        last_slot = slots_to_use[-1]
        idx = self.slots.index(last_slot)
        for extra in range(1, self.break_length_slots + 1):
            if idx + extra < len(self.slots):
                next_slot = self.slots[idx + extra]
                if timetable.at[day, next_slot] == "":
                    # Keep break as a local pacing constraint for this sheet only.
                    # Do not reserve rooms globally for break slots.
                    timetable.at[day, next_slot] = "BREAK"

        return slots_to_use


    def generate_timetable(self, courses_to_allocate, writer, sheet_name):
        timetable = pd.DataFrame("", index=self.days, columns=self.slots)
        lecturer_busy = {day: {slot: [] for slot in self.slots} for day in self.days}
        labs_scheduled = {day: False for day in self.days}
        self.course_room_map = {}

        electives = [c for c in courses_to_allocate if c.is_elective]
        non_electives = [c for c in courses_to_allocate if not c.is_elective]

        baskets = {}
        for e in electives:
            baskets.setdefault(e.basket, []).append(e)

        basket_sizes = {b: len(g) for b, g in baskets.items()}

        chosen_electives = []

        self.electives_by_sheet[sheet_name] = chosen_electives

        elective_placeholders = []

        for b in sorted(baskets.keys()):
            if b == 0:
                continue
            group = baskets[b]
            chosen = self._pick_elective_representative(b, group, sheet_name)
            if not chosen:
                continue
            chosen_electives.append((b, chosen))

            elective_course = Course(
                {
                    "Course_Code": f"Elective_{b}",
                    "Course_Title": chosen.title,
                    "Faculty": chosen.faculty,
                    "L-T-P-S-C": chosen.ltp,
                    "Semester_Half": chosen.semester_half,
                    "Elective": 0,
                }
            )
            elective_placeholders.append(elective_course)

        self.electives_by_sheet[sheet_name] = chosen_electives

        # Phase order:
        # 1) Elective placeholders first (cross-sem slot templates)
        # 2) Combined non-electives next
        # 3) Remaining regular non-electives
        combined_non_electives = [c for c in non_electives if getattr(c, "is_combined", False)]
        regular_non_electives = [c for c in non_electives if not getattr(c, "is_combined", False)]

        combined_non_electives.sort(key=lambda c: (c.P, c.L + c.T + c.S), reverse=True)
        regular_non_electives.sort(key=lambda c: (c.P, c.L + c.T + c.S), reverse=True)

        all_courses = elective_placeholders + combined_non_electives + regular_non_electives

        for course in all_courses:
            faculty, code, is_elective = course.faculty, course.code, course.code.startswith("Elective_")
            basket_id = int(code.split("_")[1]) if is_elective else None
            min_rooms = basket_sizes.get(basket_id, 1) if is_elective else 1
            is_combined = bool(getattr(course, "is_combined", False) and not is_elective)
            required_capacity_lt_t = self._required_capacity_for_course(course, is_elective, is_combined)

            # --- Lecture Scheduling ---
            # Prepare to schedule. If elective and global slots exist, use them.
            
            forced_allocations_L = []
            elective_key_L = self._elective_template_key(basket_id, "L", sheet_name) if is_elective else None
            has_template_L = bool(elective_key_L is not None and elective_key_L in self.global_elective_slots)
            if has_template_L:
                forced_allocations_L = self.global_elective_slots[elective_key_L]
            combined_key_L = self._combined_template_key(code, "L", sheet_name) if is_combined else None
            if is_combined and combined_key_L in self.global_combined_slots:
                forced_allocations_L = self.global_combined_slots[combined_key_L]

            remaining = course.L
            
            if forced_allocations_L:
                # Deterministic scheduling for this elective
                for alloc in forced_allocations_L:
                    if remaining <= 0:
                        break
                    day = alloc['day']
                    slots = alloc['slots']
                    duration = sum(self.slot_durations[s] for s in slots)
                    # Try to allocate forced slots
                    res = self._allocate_session(
                         timetable,
                         lecturer_busy,
                         labs_scheduled,
                         day,
                         faculty,
                         code,
                         duration,
                         "L",
                         is_elective,
                         sheet_name,
                         force_slots=slots,
                         min_rooms_needed=min_rooms,
                         combined_key=combined_key_L,
                         preferred_room=alloc.get("room") if is_combined else None,
                         min_capacity_needed=required_capacity_lt_t,
                    )
                    if not res and is_elective and self.relax_cross_sem_elective_block:
                        res = self._allocate_session(
                            timetable,
                            lecturer_busy,
                            labs_scheduled,
                            day,
                            faculty,
                            code,
                            duration,
                            "L",
                            is_elective,
                            sheet_name,
                            force_slots=slots,
                            min_rooms_needed=min_rooms,
                            relax_elective_block=True,
                            combined_key=combined_key_L,
                            min_capacity_needed=required_capacity_lt_t,
                        )
                    if res:
                        remaining -= duration
                        if is_combined:
                            room = self.scheduled_entries[-1].get("room", "") if self.scheduled_entries else ""
                            self._record_combined_slots(combined_key_L, day, res, room)
            
            # Standard stochastic scheduling (runs if not fully scheduled by force)
            attempts = 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_to_try = self.days.copy()
                random.shuffle(days_to_try)
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                        continue
                    alloc_dur = min(1.5, remaining)
                    
                    allocated_slots = self._allocate_session(
                        timetable,
                        lecturer_busy,
                        labs_scheduled,
                        day,
                        faculty,
                        code,
                        alloc_dur,
                        "L",
                        is_elective,
                        sheet_name,
                        min_rooms_needed=min_rooms,
                        combined_key=combined_key_L,
                        min_capacity_needed=required_capacity_lt_t,
                    )
                    if not allocated_slots and is_elective and self.relax_cross_sem_elective_block:
                        allocated_slots = self._allocate_session(
                            timetable,
                            lecturer_busy,
                            labs_scheduled,
                            day,
                            faculty,
                            code,
                            alloc_dur,
                            "L",
                            is_elective,
                            sheet_name,
                            min_rooms_needed=min_rooms,
                            relax_elective_block=True,
                            combined_key=combined_key_L,
                            min_capacity_needed=required_capacity_lt_t,
                        )
                    if allocated_slots:
                        remaining -= alloc_dur
                        if is_elective and elective_key_L and not has_template_L:
                            self.global_elective_slots.setdefault(elective_key_L, []).append({
                                'day': day,
                                'slots': allocated_slots
                            })
                        if is_combined:
                            room = self.scheduled_entries[-1].get("room", "") if self.scheduled_entries else ""
                            self._record_combined_slots(combined_key_L, day, allocated_slots, room)
                        break

            if remaining > 0:
                self.unscheduled_courses.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Lecture",
                    "remaining_hours": remaining,
                    "semester_half": course.semester_half
                })

            # --- Tutorial Scheduling ---
            forced_allocations_T = []
            elective_key_T = self._elective_template_key(basket_id, "T", sheet_name) if is_elective else None
            has_template_T = bool(elective_key_T is not None and elective_key_T in self.global_elective_slots)
            if has_template_T:
                forced_allocations_T = self.global_elective_slots[elective_key_T]
            combined_key_T = self._combined_template_key(code, "T", sheet_name) if is_combined else None
            if is_combined and combined_key_T in self.global_combined_slots:
                forced_allocations_T = self.global_combined_slots[combined_key_T]

            remaining = course.T
            if forced_allocations_T:
                 for alloc in forced_allocations_T:
                    if remaining <= 0:
                        break
                    day = alloc['day']
                    slots = alloc['slots']
                    res = self._allocate_session(
                         timetable,
                         lecturer_busy,
                         labs_scheduled,
                         day,
                         faculty,
                         code,
                         1,
                         "T",
                         is_elective,
                         sheet_name,
                         force_slots=slots,
                         min_rooms_needed=min_rooms,
                         combined_key=combined_key_T,
                         preferred_room=alloc.get("room") if is_combined else None,
                         min_capacity_needed=required_capacity_lt_t,
                    )
                    if not res and is_elective and self.relax_cross_sem_elective_block:
                        res = self._allocate_session(
                            timetable,
                            lecturer_busy,
                            labs_scheduled,
                            day,
                            faculty,
                            code,
                            1,
                            "T",
                            is_elective,
                            sheet_name,
                            force_slots=slots,
                            min_rooms_needed=min_rooms,
                            relax_elective_block=True,
                            combined_key=combined_key_T,
                            min_capacity_needed=required_capacity_lt_t,
                        )
                    if res:
                        remaining -= 1
                        if is_combined:
                            room = self.scheduled_entries[-1].get("room", "") if self.scheduled_entries else ""
                            self._record_combined_slots(combined_key_T, day, res, room)

            attempts = 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_to_try = self.days.copy()
                random.shuffle(days_to_try)
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                        continue
                    
                    allocated_slots = self._allocate_session(
                        timetable,
                        lecturer_busy,
                        labs_scheduled,
                        day,
                        faculty,
                        code,
                        1,
                        "T",
                        is_elective,
                        sheet_name,
                        min_rooms_needed=min_rooms,
                        combined_key=combined_key_T,
                        min_capacity_needed=required_capacity_lt_t,
                    )
                    if not allocated_slots and is_elective and self.relax_cross_sem_elective_block:
                        allocated_slots = self._allocate_session(
                            timetable,
                            lecturer_busy,
                            labs_scheduled,
                            day,
                            faculty,
                            code,
                            1,
                            "T",
                            is_elective,
                            sheet_name,
                            min_rooms_needed=min_rooms,
                            relax_elective_block=True,
                            combined_key=combined_key_T,
                            min_capacity_needed=required_capacity_lt_t,
                        )
                    if allocated_slots:
                        remaining -= 1
                        if is_elective and elective_key_T and not has_template_T:
                            self.global_elective_slots.setdefault(elective_key_T, []).append({
                                'day': day,
                                'slots': allocated_slots
                            })
                        if is_combined:
                            room = self.scheduled_entries[-1].get("room", "") if self.scheduled_entries else ""
                            self._record_combined_slots(combined_key_T, day, allocated_slots, room)
                        break

            if remaining > 0:
                self.unscheduled_courses.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Tutorial",
                    "remaining_hours": remaining,
                    "semester_half": course.semester_half
                })

            # --- Lab Scheduling ---
            forced_allocations_P = []
            elective_key_P = self._elective_template_key(basket_id, "P", sheet_name) if is_elective else None
            has_template_P = bool(elective_key_P is not None and elective_key_P in self.global_elective_slots)
            if has_template_P:
                forced_allocations_P = self.global_elective_slots[elective_key_P]
            combined_key_P = self._combined_template_key(code, "P", sheet_name) if is_combined else None
            if is_combined and combined_key_P in self.global_combined_slots:
                forced_allocations_P = self.global_combined_slots[combined_key_P]

            remaining = course.P
            if forced_allocations_P:
                 for alloc in forced_allocations_P:
                    if remaining <= 0:
                        break
                    day = alloc['day']
                    slots = alloc['slots']
                    duration = sum(self.slot_durations[s] for s in slots)
                    res = self._allocate_session(
                         timetable,
                         lecturer_busy,
                         labs_scheduled,
                         day,
                         faculty,
                         code,
                         duration,
                         "P",
                         is_elective,
                         sheet_name,
                         force_slots=slots,
                         min_rooms_needed=min_rooms,
                         combined_key=combined_key_P,
                         preferred_room=alloc.get("room") if is_combined else None,
                         min_capacity_needed=None,
                    )
                    if not res and is_elective and self.relax_cross_sem_elective_block:
                        res = self._allocate_session(
                            timetable,
                            lecturer_busy,
                            labs_scheduled,
                            day,
                            faculty,
                            code,
                            duration,
                            "P",
                            is_elective,
                            sheet_name,
                            force_slots=slots,
                            min_rooms_needed=min_rooms,
                            relax_elective_block=True,
                            combined_key=combined_key_P,
                            min_capacity_needed=None,
                        )
                    if res:
                        remaining -= duration
                        if is_combined:
                            room = self.scheduled_entries[-1].get("room", "") if self.scheduled_entries else ""
                            self._record_combined_slots(combined_key_P, day, res, room)
            
            attempts = 0
            while remaining > 0 and attempts < self.MAX_ATTEMPTS:
                attempts += 1
                days_without_labs = [d for d in self.days if not labs_scheduled[d]]
                days_to_try = days_without_labs.copy()
                random.shuffle(days_to_try)
                for day in days_to_try:
                    if remaining <= 0 or (faculty and faculty in lecturer_busy[day]):
                        continue
                    alloc_dur = min(2, remaining) if remaining >= 2 else remaining
                    
                    allocated_slots = self._allocate_session(
                        timetable,
                        lecturer_busy,
                        labs_scheduled,
                        day,
                        faculty,
                        code,
                        alloc_dur,
                        "P",
                        is_elective,
                        sheet_name,
                        min_rooms_needed=min_rooms,
                        combined_key=combined_key_P,
                        min_capacity_needed=None,
                    )
                    if not allocated_slots and is_elective and self.relax_cross_sem_elective_block:
                        allocated_slots = self._allocate_session(
                            timetable,
                            lecturer_busy,
                            labs_scheduled,
                            day,
                            faculty,
                            code,
                            alloc_dur,
                            "P",
                            is_elective,
                            sheet_name,
                            min_rooms_needed=min_rooms,
                            relax_elective_block=True,
                            combined_key=combined_key_P,
                            min_capacity_needed=None,
                        )
                    if allocated_slots:
                        remaining -= alloc_dur
                        if is_elective and elective_key_P and not has_template_P:
                            self.global_elective_slots.setdefault(elective_key_P, []).append({
                                'day': day,
                                'slots': allocated_slots
                            })
                        if is_combined:
                            room = self.scheduled_entries[-1].get("room", "") if self.scheduled_entries else ""
                            self._record_combined_slots(combined_key_P, day, allocated_slots, room)
                        break

            if remaining > 0:
                self.unscheduled_courses.append({
                    "sheet": sheet_name,
                    "course_code": code,
                    "course_title": course.title,
                    "faculty": faculty,
                    "type": "Lab",
                    "remaining_hours": remaining,
                    "semester_half": course.semester_half
                })

        for day in self.days:
            for slot in self.excluded_slots:
                if slot in timetable.columns:
                    timetable.at[day, slot] = ""

        timetable.to_excel(writer, sheet_name=sheet_name, index=True)
        print(f"Saved timetable to sheet '{sheet_name}'")


    def _compute_elective_room_assignments_legally(self, sheet_name):
        electives_representatives = self.electives_by_sheet.get(sheet_name, [])
        if not electives_representatives:
            self.elective_room_assignment[sheet_name] = {}
            return

        active_baskets = set(b for b, _ in electives_representatives)
        assigned = {}
        room_usage = self._sheet_scoped_usage(self.global_room_usage, sheet_name)
        elective_room_usage = self._sheet_scoped_usage(self.global_elective_room_usage, sheet_name)
        
        # Local usage tracker for electives in this sheet: day -> slot -> set of rooms used
        local_room_usage = {} 

        def is_display_room_free(day, slot, room, owner_key, check_non_elective_usage=True):
            if check_non_elective_usage and room in room_usage.get(day, {}).get(slot, []):
                return False
            used_by = elective_room_usage.get(day, {}).get(slot, {}).get(room)
            if used_by is not None and used_by != owner_key:
                return False
            if room in local_room_usage.get(day, {}).get(slot, set()):
                return False
            return True

        def reserve_display_room(day, slot, room, owner_key):
            local_room_usage.setdefault(day, {}).setdefault(slot, set()).add(room)
            elective_room_usage.setdefault(day, {}).setdefault(slot, {})[room] = owner_key

        for basket in sorted(active_baskets):
            basket_code = f"Elective_{basket}"
            # Get slots for this basket
            lecture_slots = []
            lab_slots = []
            for ent in self.scheduled_entries:
                if ent["sheet"] == sheet_name and ent["code"] == basket_code:
                    if "(Lab" in ent["display"]: 
                         lab_slots.append((ent["day"], ent["slot"]))
                    else:
                         lecture_slots.append((ent["day"], ent["slot"]))
            
            # Prioritize lecture slots for room assignment (Classrooms). 
            # If no lecture slots (pure lab), use lab slots.
            # This prevents Lab slots (which happen in Labs) from blocking Classrooms for the Lecture component.
            raw_slots = lecture_slots if lecture_slots else lab_slots
            basket_slots = sorted(list(set(raw_slots)))
            
            # Even if no slots found (e.g. unscheduled), we proceed (will act as 0 duration) 
            # but room assignment requires slots to check conflicts. 
            # If no slots, room assignment is arbitrary (free).
            
            basket_electives = [
                c
                for c in self.courses
                if c.is_elective and c.basket == basket and self._course_in_sheet_half(c, sheet_name)
            ]
            # For electives that have lecture slots, restrict rooms to classrooms so lectures never land in labs.
            # Pure-lab electives still prefer labs first but can spill into classrooms if needed.
            if lecture_slots:
                candidate_rooms = [
                    r
                    for r in sorted(self.classrooms)
                    if self._room_allowed_for_course(r, is_compulsory=False)
                ]
            else:
                candidate_rooms = sorted(self.labs) + [
                    r
                    for r in sorted(self.classrooms)
                    if self._room_allowed_for_course(r, is_compulsory=False)
                ]
            
            for elective in basket_electives:
                key = f"Elective_{basket}||{elective.title}"
                template_key = (self.semester_group, sheet_name, basket, elective.title.strip().lower())
                preferred_room = self.global_elective_room_templates.get(template_key)
                
                # OPTIMIZATION: Try to find ONE stable room for all slots first
                stable_room = None
                
                # Check all rooms for stability
                # Preference order is driven by candidate_rooms computed above.
                search_order = candidate_rooms

                if preferred_room and self._room_allowed_for_course(preferred_room, is_compulsory=False):
                    preferred_free = True
                    for day, slot in basket_slots:
                        if not is_display_room_free(
                            day,
                            slot,
                            preferred_room,
                            template_key,
                            check_non_elective_usage=False,
                        ):
                            preferred_free = False
                            break
                    if preferred_free:
                        stable_room = preferred_room

                if not stable_room:
                    for r in search_order: 
                         is_universally_free = True
                         for day, slot in basket_slots:
                             if not is_display_room_free(day, slot, r, template_key):
                                 is_universally_free = False
                                 break
                         
                         if is_universally_free:
                             stable_room = r
                             break
                
                used_rooms_for_this_elective = []
                
                if stable_room:
                    # Perfect! We found one room for everything.
                    for day, slot in basket_slots:
                        reserve_display_room(day, slot, stable_room, template_key)
                    used_rooms_for_this_elective.append(stable_room)
                    self.global_elective_room_templates[template_key] = stable_room
                else:
                    # Fallback: Assign room PER SLOT
                    room_assignments = [] # List of (day, room)
                    
                    last_used_room = None
                    
                    for day, slot in basket_slots:
                        # Reuse last room if possible
                        candidates_ordered = []
                        if last_used_room:
                            candidates_ordered.append(last_used_room)
                            candidates_ordered.extend([r for r in candidate_rooms if r != last_used_room])
                        else:
                            candidates_ordered = candidate_rooms

                        chosen_for_slot = None
                        
                        for r in candidates_ordered:
                            if not is_display_room_free(day, slot, r, template_key):
                                continue
                            chosen_for_slot = r
                            break
                        
                        if not chosen_for_slot:
                             for r in sorted(self.all_rooms):
                                 if not self._room_allowed_for_course(r, is_compulsory=False):
                                     continue
                                 if is_display_room_free(day, slot, r, template_key):
                                     chosen_for_slot = r
                                     break
                        
                        if chosen_for_slot:
                            reserve_display_room(day, slot, chosen_for_slot, template_key)
                            room_assignments.append((day, chosen_for_slot))
                            last_used_room = chosen_for_slot
                    if room_assignments:
                        unique_rooms = {r for _, r in room_assignments}
                        if len(unique_rooms) == 1:
                            self.global_elective_room_templates[template_key] = next(iter(unique_rooms))

                # Format the room string
                if stable_room:
                    assigned[key] = stable_room
                else:
                    # Group by room -> days
                    room_to_days = {}
                    for day, room in room_assignments:
                        room_to_days.setdefault(room, set()).add(day)
                    
                    # Sort rooms by first appearance day (approx) or name
                    # Let's sort by room name for consistency or by schedule order? 
                    # Schedule order is better but complex. Room name is fine.
                    
                    parts = []
                    # Pre-defined abbreviations
                    day_abbr = {"Monday": "Mon", "Tuesday": "Tue", "Wednesday": "Wed", "Thursday": "Thu", "Friday": "Fri"}
                    
                    # Sort logic: sort by first day index found for that room
                    def sort_key(item):
                        r, days = item
                        # Min index of days
                        indices = [self.days.index(d) for d in days if d in self.days]
                        return min(indices) if indices else 99
                    
                    sorted_items = sorted(room_to_days.items(), key=sort_key)
                    
                    for room, days in sorted_items:
                        # Sort days
                        sorted_days = sorted(list(days), key=lambda d: self.days.index(d) if d in self.days else 99)
                        d_strs = [day_abbr.get(d, d[:3]) for d in sorted_days]
                        parts.append(f"{room} ({','.join(d_strs)})")
                    
                    assigned[key] = ", ".join(parts)

        self.elective_room_assignment[sheet_name] = assigned

    def format_student_timetable_with_legend(self, filename):
        wb = load_workbook(filename)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        color_map = {}
        palette = ["FFC7CE", "C6EFCE", "FFEB9C", "BDD7EE", "D9EAD3", "F4CCCC",
                "D9D2E9", "FCE5CD", "C9DAF8", "EAD1DC"]
        color_index = 0

        for sheet_name in wb.sheetnames:
            self._compute_elective_room_assignments_legally(sheet_name)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in range(2, ws.max_row + 1):
                start_col = 2
                while start_col <= ws.max_column:
                    cell = ws.cell(row=row, column=start_col)
                    val = str(cell.value).strip() if cell.value is not None else ""

                    if val and val not in ["FREE", "BREAK"]:
                        raw_code = val.split(" ")[0]
                        code = raw_code.rstrip("T")
                        if code not in color_map:
                            color_map[code] = palette[color_index % len(palette)]
                            color_index += 1
                        fill = PatternFill(start_color=color_map[code], end_color=color_map[code], fill_type="solid")
                        cell.fill = fill

                        merge_count = 0
                        for col in range(start_col + 1, ws.max_column + 1):
                            next_cell = ws.cell(row=row, column=col)
                            if next_cell.value == cell.value:
                                next_cell.fill = fill
                                merge_count += 1
                            else:
                                break
                        if merge_count > 0:
                            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + merge_count)

                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        for col_idx in range(start_col, start_col + merge_count + 1):
                            ws.cell(row=row, column=col_idx).border = thin_border
                        start_col += merge_count + 1
                    elif val == "BREAK":
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        start_col += 1
                    else:
                        cell.border = thin_border
                        start_col += 1

            start_row = ws.max_row + 3
            headers = ["S.No", "Course Code", "Course Title", "L-T-P-S-C", "Faculty", "Color"]
            for idx, header in enumerate(headers, start=2):
                ws.cell(start_row, idx, header).border = thin_border
                ws.cell(start_row, idx).alignment = Alignment(horizontal="center", vertical="center")


            i = 1
            for code in color_map:
                if code.startswith("Elective_"):
                    continue
                ws.cell(start_row + i, 2, i).border = thin_border
                ws.cell(start_row + i, 3, code).border = thin_border
                course_name = next((c.title for c in self.courses if c.code == code), code)
                faculty = next((c.faculty for c in self.courses if c.code == code), "")
                ltpsc = next((c.ltp for c in self.courses if c.code == code), "")
                ws.cell(start_row + i, 4, course_name).border = thin_border
                ws.cell(start_row + i, 5, ltpsc).border = thin_border
                ws.cell(start_row + i, 5).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(start_row + i, 6, faculty).border = thin_border
                ws.cell(start_row + i, 7, "").fill = PatternFill(
                    start_color=color_map[code],
                    end_color=color_map[code],
                    fill_type="solid"
                )
                ws.cell(start_row + i, 7).border = thin_border

                i += 1

            electives_header_row = start_row + i + 2
            e_headers = ["S.No", "Elective Basket", "Elective Title", "Faculty", "Room", "Color"]
            for idx, header in enumerate(e_headers, start=2):
                ws.cell(electives_header_row, idx, header).border = thin_border
                ws.cell(electives_header_row, idx).alignment = Alignment(horizontal="center", vertical="center")

            chosen_by_basket = {b: e for (b, e) in self.electives_by_sheet.get(sheet_name, [])}
            row_ctr = 1

            for basket in sorted(chosen_by_basket.keys()):
                elective_code = f"Elective_{basket}"
                chosen_e = chosen_by_basket[basket]

                all_electives = [
                    c
                    for c in self.courses
                    if c.is_elective and c.basket == basket and self._course_in_sheet_half(c, sheet_name)
                ]

                for idx_e, e in enumerate(all_electives):
                    ws.cell(electives_header_row + row_ctr, 2, row_ctr).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 3, elective_code).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 4, e.title).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 5, e.faculty).border = thin_border
                    
                    key = f"{elective_code}||{e.title}"
                    room = self.elective_room_assignment.get(sheet_name, {}).get(key, "")
                    ws.cell(electives_header_row + row_ctr, 6, room).border = thin_border
                    ws.cell(electives_header_row + row_ctr, 6).alignment = Alignment(horizontal="center", vertical="center")

                    ws.cell(electives_header_row + row_ctr, 7, "").fill = PatternFill(
                        start_color=color_map.get(elective_code, "FFFFFF"),
                        end_color=color_map.get(elective_code, "FFFFFF"),
                        fill_type="solid"
                    )
                    ws.cell(electives_header_row + row_ctr, 7).border = thin_border

                    row_ctr += 1

            timetable_max_row = len(self.days) + 1
            timetable_max_col = ws.max_column
            for row in ws.iter_rows(min_row=2, max_row=timetable_max_row, min_col=2, max_col=timetable_max_col):
                for cell in row:
                    cell.border = thin_border
            ws.freeze_panes = "B2"


            for col in ws.columns:
                max_length = 0
                try:
                    column_letter = col[0].column_letter
                except Exception:
                    continue
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = max_length + 2

        wb.save(filename)
        print(f"Formatted student timetable saved in {filename}")



    def _generate_faculty_workbook(self, faculty_filename):
        faculty_set = set()
        for c in self.courses:
            raw = c.faculty
            if raw:
                for p in [x.strip() for x in raw.split("/") if x.strip()]:
                    faculty_set.add(p)
        for ent in self.scheduled_entries:
            fstr = ent.get("faculty")
            if fstr:
                for p in [x.strip() for x in fstr.split("/") if x.strip()]:
                    faculty_set.add(p)

        faculty_tables = {}
        for f in faculty_set:
            faculty_tables[f] = {
                "First_Half": pd.DataFrame("    ", index=self.days, columns=self.slots),
                "Second_Half": pd.DataFrame("    ", index=self.days, columns=self.slots)
            }

        for ent in self.scheduled_entries:
            day = ent["day"]
            slot = ent["slot"]
            display = ent["display"]
            sheet = ent["sheet"]
            if ent.get("faculty"):
                faculties = [p.strip() for p in ent["faculty"].split("/") if p.strip()]
            else:
                matched = [c for c in self.courses if c.code == ent.get("code")]
                faculties = []
                for m in matched:
                    faculties.extend([p.strip() for p in m.faculty.split("/") if p.strip()])

            for f in faculties:
                faculty_tables[f][sheet].at[day, slot] = display

        with pd.ExcelWriter(faculty_filename, engine="openpyxl") as writer:
            for f in sorted(faculty_tables.keys()):
                safe = f[:31]
                df_first = faculty_tables[f]["First_Half"]
                df_second = faculty_tables[f]["Second_Half"]

                combined = pd.DataFrame()
                combined[" "] = [""] * (len(self.days) + 2)
                temp = df_first.reset_index()
                temp.columns = ["Day"] + list(df_first.columns)
                combined_first = temp

                spacer = pd.DataFrame([[""] * temp.shape[1]], columns=temp.columns)

                temp2 = df_second.reset_index()
                temp2.columns = ["Day"] + list(df_second.columns)
                combined_second = temp2

                final = pd.concat([combined_first, spacer, combined_second], ignore_index=True)
                final.to_excel(writer, sheet_name=safe, index=False)

        wb = load_workbook(faculty_filename)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        palette = ["FFC7CE","C6EFCE","FFEB9C","BDD7EE","D9EAD3","F4CCCC","D9D2E9","FCE5CD","C9DAF8","EAD1DC"]
        color_map = {}
        color_index = 0

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in range(2, ws.max_row + 1):
                start_col = 2
                while start_col <= ws.max_column:
                    cell = ws.cell(row=row, column=start_col)
                    if cell.value and str(cell.value).strip() not in ["FREE", ""]:
                        raw_code = str(cell.value).split(" ")[0].rstrip("T")
                        if raw_code not in color_map:
                            color_map[raw_code] = palette[color_index % len(palette)]
                            color_index += 1
                        fill = PatternFill(start_color=color_map[raw_code], end_color=color_map[raw_code], fill_type="solid")
                        cell.fill = fill
                        merge_count = 0
                        for col in range(start_col + 1, ws.max_column + 1):
                            next_cell = ws.cell(row=row, column=col)
                            if next_cell.value == cell.value:
                                next_cell.fill = fill
                                merge_count += 1
                            else:
                                break
                        if merge_count > 0:
                            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + merge_count)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        for col_idx in range(start_col, start_col + merge_count + 1):
                            ws.cell(row=row, column=col_idx).border = thin
                        start_col += merge_count + 1
                    else:
                        cell.border = thin
                        start_col += 1

                        
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin

            ws.freeze_panes = "B2"


            for col in ws.columns:
                max_length = 0
                try:
                    column = col[0].column_letter
                except:
                    continue
                for cell in col:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2

        wb.save(faculty_filename)
        print(f"Saved faculty timetables to {faculty_filename}")

    def run_all_outputs(self, dept_name_prefix="CSE", student_filename=None, faculty_filename="faculty_timetable.xlsx"):
        
        if not student_filename:
            student_filename = f"{dept_name_prefix}_timetable.xlsx"

        
        self.scheduled_entries = []
        self.electives_by_sheet = {}
        self.elective_room_assignment = {}

        with pd.ExcelWriter(student_filename, engine="openpyxl") as writer:
           
            self.generate_timetable([c for c in self.courses if c.semester_half in ["1", "0"]], writer, "First_Half")
            self.generate_timetable([c for c in self.courses if c.semester_half in ["2", "0"]], writer, "Second_Half")

        if self.unscheduled_courses:
            unsched_file = f"{dept_name_prefix}_unscheduled_courses.xlsx"
            df_unsched = pd.DataFrame(self.unscheduled_courses)
            df_unsched.to_excel(unsched_file, index=False)
            print(f"Some courses couldn't be scheduled. See '{unsched_file}' for details.")
            print(df_unsched.to_string(index=False))
        wb = load_workbook(student_filename)
        for default in ["Sheet", "Sheet1"]:
            if default in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb[default])
        wb.save(student_filename)

        
        for sheet in wb.sheetnames:
            self._compute_elective_room_assignments_legally(sheet)

        
        self.format_student_timetable_with_legend(student_filename)

        
        self._generate_faculty_workbook(faculty_filename)


def _resolve_combined_cluster_from_dept(dept_name):
    dept_name = str(dept_name).strip()
    dept_prefix = dept_name.split("-")[0].strip().upper() if dept_name else ""
    cluster_groups = (
        {"CSE"},
        {"DSAI", "ECE"},
    )
    for group in cluster_groups:
        if dept_prefix in group:
            return "+".join(sorted(group))
    return dept_prefix if dept_prefix else dept_name.upper()


def _safe_students_count(value):
    try:
        return max(0, int(float(str(value).strip())))
    except Exception:
        return 0


def _is_truthy_flag(value):
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _is_elective_row(row):
    elective_raw = row.get("Elective", 0)
    elective_str = str(elective_raw).strip().lower()
    elective_num = 0.0
    try:
        elective_num = float(elective_str) if elective_str else 0.0
    except Exception:
        elective_num = 1.0 if elective_str in {"true", "yes", "y"} else 0.0
    basket_val = _safe_students_count(row.get("basket", 0))
    return (elective_num > 0) or (basket_val > 0)


def _build_global_combined_strength(departments):
    totals = {}
    for dept_name, course_file in departments.items():
        match = re.search(r"\d+", str(dept_name))
        semester_group = match.group(0) if match else "UNKNOWN"
        cluster_id = _resolve_combined_cluster_from_dept(dept_name)
        df = pd.read_csv(course_file)
        for _, row in df.iterrows():
            if not _is_truthy_flag(row.get("is_combined", row.get("Is_Combined", 0))):
                continue
            if _is_elective_row(row):
                continue
            code = str(row.get("Course_Code", "")).strip().upper()
            if not code:
                continue
            students = _safe_students_count(row.get("Students", 0))
            key = (semester_group, cluster_id, code)
            totals[key] = totals.get(key, 0) + students
    return totals


if __name__ == "__main__":
   
    departments = {
        "CSE-3-A": "data/coursesCSEA-III.csv",
        "CSE-3-B": "data/coursesCSEB-III.csv",
        "CSE-1-A": "data/coursesCSEA-I.csv",
        "CSE-1-B": "data/coursesCSEB-I.csv",
        "CSE-5-A": "data/coursesCSEA-V.csv",
        "CSE-5-B": "data/coursesCSEB-V.csv",
        "7-SEM": "data/courses7.csv",
        "DSAI-3": "data/coursesDSAI-III.csv",
        "ECE-3": "data/coursesECE-III.csv",
        "DSAI-1": "data/coursesDSAI-I.csv",
        "ECE-1": "data/coursesECE-I.csv",
        "DSAI-5": "data/coursesDSAI-V.csv",
        "ECE-5": "data/coursesECE-V.csv",
    }
    rooms_file = "data/rooms.csv"
    slots_file = "data/timeslots.csv"
    global_room_usage = {}
    combined_faculty_filename = "faculty_timetable.xlsx"
    
    global_elective_slots = {}
    global_elective_slot_usage = {}
    global_elective_room_templates = {}
    global_elective_room_usage = {}
    global_elective_representatives = {}
    global_combined_slots = {}
    global_combined_room_usage = {}
    global_combined_strength = _build_global_combined_strength(departments)
    global_c004_reserved_slots = {}
    all_scheduled_entries = []

    for dept_name, course_file in departments.items():
        print(f"\nGenerating student timetable for {dept_name}...")
        scheduler = Scheduler(
            slots_file,
            course_file,
            rooms_file,
            global_room_usage,
            global_elective_slots,
            dept_name=dept_name,
            global_elective_slot_usage=global_elective_slot_usage,
            global_elective_room_templates=global_elective_room_templates,
            global_elective_room_usage=global_elective_room_usage,
            global_elective_representatives=global_elective_representatives,
            global_combined_slots=global_combined_slots,
            global_combined_room_usage=global_combined_room_usage,
            global_combined_strength=global_combined_strength,
            global_c004_reserved_slots=global_c004_reserved_slots,
        )
        student_file = f"{dept_name}_timetable.xlsx"
        scheduler.run_all_outputs(dept_name_prefix=dept_name, student_filename=student_file, faculty_filename=combined_faculty_filename)

        all_scheduled_entries.extend(scheduler.scheduled_entries)
        for k, v in scheduler.course_room_map.items():    
            global_room_usage.setdefault("MAPPING", {})[k] = v    
    combined_courses = []
    for dept_name, course_file in departments.items():
        df = pd.read_csv(course_file)
        for _, row in df.iterrows():
            combined_courses.append(Course(row))
    helper = Scheduler(
        slots_file,
        departments[list(departments.keys())[0]],
        rooms_file,
        global_room_usage,
        global_elective_slots,
        global_elective_slot_usage=global_elective_slot_usage,
        global_elective_room_templates=global_elective_room_templates,
        global_elective_room_usage=global_elective_room_usage,
        global_elective_representatives=global_elective_representatives,
        global_combined_slots=global_combined_slots,
        global_combined_room_usage=global_combined_room_usage,
        global_combined_strength=global_combined_strength,
        global_c004_reserved_slots=global_c004_reserved_slots,
    )
    helper.courses = combined_courses
    helper.scheduled_entries = all_scheduled_entries 
    helper._generate_faculty_workbook(combined_faculty_filename)
    print("\nAll done. Student timetables and combined faculty timetable generated.")
