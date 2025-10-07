import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
import csv
import json
import os
import traceback

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
LECTURE_DURATION = 3  # 1.5 hours = 3 slots (30 mins each)
LAB_DURATION = 4      # 2 hours = 4 slots (30 mins each)
TUTORIAL_DURATION = 2  # 1 hour = 2 slots (30 mins each)
SELF_STUDY_DURATION = 2  # 1 hour = 2 slots (30 mins each)
BREAK_DURATION = 1    # 30 mins = 1 slot

# Lunch break parameters
LUNCH_WINDOW_START = time(12, 30)  # Lunch breaks can start from 12:30
LUNCH_WINDOW_END = time(14, 0)    # Last lunch break must end by 14:00 
LUNCH_DURATION = 60              # Each semester gets 45 min lunch

# Color palette for different courses (vibrant colors)
COLOR_PALETTE = [
    "FF5733", "33FF57", "3357FF", "FF33A8", "33FFF7", 
    "F7FF33", "FF33F7", "33F7FF", "FFB533", "B533FF",
    "33FFB5", "FF5F33", "335FFF", "B5FF33", "FF33B5"
]

# Load room data
try:
    rooms_df = pd.read_csv('rooms.csv')
    # Create separate lists for lecture rooms and lab rooms
    lecture_rooms = rooms_df[rooms_df['type'] == 'LECTURE_ROOM']['roomNumber'].tolist()
    computer_lab_rooms = rooms_df[rooms_df['type'] == 'COMPUTER_LAB']['roomNumber'].tolist()
    large_rooms = rooms_df[rooms_df['type'] == 'SEATER_120']['roomNumber'].tolist()
    
    # Add error handling for room types
    if not lecture_rooms:
        print("Warning: No LECTURE_ROOM type rooms found in rooms.csv")
    if not computer_lab_rooms:
        print("Warning: No COMPUTER_LAB type rooms found in rooms.csv")
    if not large_rooms:
        print("Warning: No SEATER_120 type rooms found in rooms.csv")
except FileNotFoundError:
    print("Error: File 'rooms.csv' not found in the current directory")
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []
except Exception as e:
    print(f"Error loading rooms.csv: {e}")
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []

def generate_course_color():
    """Generate unique colors for courses from the palette or random if needed"""
    for color in COLOR_PALETTE:
        yield color
    
    # If we run out of predefined colors, generate random ones
    while True:
        r = format(random.randint(180, 255), '02x')
        g = format(random.randint(180, 255), '02x')
        b = format(random.randint(180, 255), '02x')
        yield f"{r}{g}{b}"

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=30)
        
        # Keep all time slots but we'll mark break times later
        slots.append((current, next_time.time()))
        current_time = next_time
    
    return slots

# Load data from CSV
try:
    df = pd.read_csv('combined.csv')
except FileNotFoundError:
    print("Error: File 'combined.csv' not found in the current directory")
    exit()

def is_break_time(slot):
    """Check if a time slot falls within break times"""
    start, end = slot
    # Morning break: 10:30-11:00 (15 minutes)
    morning_break = (time(10, 30) <= start < time(10, 45))
    # Lunch break: 13:00-14:00 (45 minutes)
    lunch_break = (time(13, 0) <= start < time(13, 45))
    # Inter-class break (5 minutes)
    # This is handled in the scheduling logic, not here
    return morning_break or lunch_break

def check_professor_availability(professor_schedule, faculty, day, start_slot, duration, activity_type):
    """Check if a professor can be scheduled for a new class considering REQ-10"""
    # If this is a lab scheduled after a lecture/tutorial, we can allow it without time gap
    # We'd need to know the previous activity type, which isn't implemented yet
    
    # Check for any existing slots for this professor on this day
    existing_slots = sorted(list(professor_schedule[faculty][day]))
    if not existing_slots:
        return True  # No other classes on this day
        
    # Calculate end slot of proposed new class
    end_slot = start_slot + duration - 1
    
    # Check the minimum time difference between existing and new slots
    MIN_GAP_SLOTS = 6  # 3 hours = 6 half-hour slots
    
    for slot in existing_slots:
        # If existing slot is within or adjacent to requested slot range
        if start_slot <= slot <= end_slot:
            return False  # Direct conflict
            
        # Check if gap between classes is sufficient
        if slot < start_slot and start_slot - slot < MIN_GAP_SLOTS:
            # Existing class ends too close to new class
            return False
            
        if slot > end_slot and slot - end_slot < MIN_GAP_SLOTS:
            # New class ends too close to existing class
            return False
    
    return True

def check_professor_constraint(professor_schedule, faculty, day, start_slot, duration, timetable, time_slots):
    """Check if a professor can be scheduled for a new class considering REQ-10"""
    # If professor has no classes that day, constraint is satisfied
    if not professor_schedule[faculty][day]:
        return True
    
    # Get the new class's time range
    new_class_start_time = time_slots[start_slot][0]
    new_class_end_time = time_slots[start_slot + duration - 1][1]
    
    # Create datetime objects for today to allow time subtraction
    today = datetime.today().date()
    new_start_datetime = datetime.combine(today, new_class_start_time)
    new_end_datetime = datetime.combine(today, new_class_end_time)
    
    # Check each existing class for this professor on this day
    for existing_slot in professor_schedule[faculty][day]:
        # Get information about the existing class
        existing_class_type = timetable[day][existing_slot]['type']
        
        # Skip slots that don't have class type (might be continuation slots)
        if existing_class_type is None:
            continue
        
        # Get existing class time
        existing_class_start_time = time_slots[existing_slot][0]
        existing_class_end_time = None
        
        # Find the end time by looking for the last slot of this class
        for i in range(existing_slot, len(time_slots)):
            if i in professor_schedule[faculty][day] and timetable[day][i]['type'] is not None:
                existing_class_end_time = time_slots[i][1]
            else:
                break
        
        # Skip if we couldn't determine the end time
        if existing_class_end_time is None:
            continue
        
        # Convert to datetime objects for comparisons
        existing_start_datetime = datetime.combine(today, existing_class_start_time)
        existing_end_datetime = datetime.combine(today, existing_class_end_time)
        
        # Special case: Lab can be scheduled after lecture/tutorial
        if (new_start_datetime == existing_end_datetime and 
            existing_class_type in ['LEC', 'TUT'] and duration == LAB_DURATION):
            continue
        
        if (existing_start_datetime == new_end_datetime and 
            timetable[day][start_slot]['type'] == 'LAB' and existing_class_type in ['LEC', 'TUT']):
            continue
        
        # Calculate time difference in hours
        time_diff_hours = abs((new_start_datetime - existing_start_datetime).total_seconds() / 3600)
        
        # Check if classes are consecutive (not allowed unless it's the special case)
        if new_start_datetime == existing_end_datetime or existing_start_datetime == new_end_datetime:
            return False
        
        # Check if the time difference is less than 3 hours
        if time_diff_hours < 3:
            return False
    
    return True

def generate_all_timetables():
    """Generate a single timetable for all departments and semesters with basket course support"""
    global TIME_SLOTS
    initialize_time_slots()  # Initialize time slots before using
    
    # Load configuration and required data
    rooms = load_rooms()
    batch_info = load_batch_data()
    
    # Create a single workbook for all timetables
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create an overview sheet
    overview_sheet = wb.create_sheet(title="Overview")
    overview_sheet.append(["Combined Timetable for All Departments and Semesters"])
    overview_sheet.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview_sheet.append([])
    overview_sheet.append(["Department", "Semester", "Sheet Name"])
    
    # Track professor assignments and unscheduled components
    professor_schedule = {}
    unscheduled_components = set()
    
    # Color palette for subjects (vibrant colors)
    subject_colors = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1", 
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]

    # Add specific colors for basket groups
    basket_group_colors = {
        'B1': "FFA07A",  # Light salmon
        'B2': "98FB98",  # Pale green  
        'B3': "87CEFA",  # Light sky blue
        'B4': "FFD700",  # Gold
        'B5': "DA70D6",  # Orchid
        'B6': "20B2AA",  # Light sea green
        'B7': "FF6347",  # Tomato
        'B8': "8A2BE2",  # Blue violet
        'B9': "32CD32"   # Lime green
    }
    
    # Add a list to track self-study only courses
    self_study_courses = []
    
    # Get all unique semester numbers for lunch breaks
    all_semesters = sorted(set(int(str(sem)[0]) for sem in df['Semester'].unique()))
    
    # Calculate lunch breaks dynamically
    lunch_breaks = calculate_lunch_breaks(all_semesters)

    # Process each department
    row_index = 5  # Starting row for overview links
    
    for department in df['Department'].unique():
        # Track assigned faculty for courses
        course_faculty_assignments = {}
        
        # Process all semesters for this department
        for semester in df[df['Department'] == department]['Semester'].unique():
            # Filter out courses marked as not to be scheduled
            courses = df[(df['Department'] == department) & 
                        (df['Semester'] == semester) &
                        ((df['Schedule'].fillna('Yes').str.upper() == 'YES') | 
                         (df['Schedule'].isna()))].copy()
            
            if courses.empty:
                continue

            # First handle lab scheduling as a separate pass
            lab_courses = courses[courses['P'] > 0].copy()
            lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
            lab_courses = lab_courses.sort_values('priority', ascending=False)

            # Handle remaining courses after labs
            non_lab_courses = courses[courses['P'] == 0].copy()
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            # Combine sorted courses with labs first
            courses = pd.concat([lab_courses, non_lab_courses])

            # Get section info
            dept_info = batch_info.get((department, semester))
            num_sections = dept_info['num_sections'] if dept_info else 1

            # First identify self-study only courses
            for _, course in courses.iterrows():
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                p = int(course['P']) if pd.notna(course['P']) else 0
                s = int(course['S']) if pd.notna(course['S']) else 0
                
                if s > 0 and l == 0 and t == 0 and p == 0:
                    self_study_courses.append({
                        'code': str(course['Course Code']),
                        'name': str(course['Course Name']),
                        'faculty': str(course['Faculty']),
                        'department': department,
                        'semester': semester
                    })

            # Process each section
            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                
                # Add to overview sheet
                overview_sheet.cell(row=row_index, column=1, value=department)
                overview_sheet.cell(row=row_index, column=2, value=str(semester))
                overview_sheet.cell(row=row_index, column=3, value=section_title)
                row_index += 1
                
                # Initialize timetable structure
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
                
                # Create a mapping for subject colors
                subject_color_map = {}
                course_faculty_map = {}  # For legend
                color_idx = 0
                
                # Assign colors to each unique subject
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code and code != 'nan':
                        if is_basket_course(code):
                            basket_group = get_basket_group(code)
                            # Use predefined basket group color
                            subject_color_map[code] = basket_group_colors.get(basket_group, subject_colors[color_idx % len(subject_colors)])
                        else:
                            subject_color_map[code] = subject_colors[color_idx % len(subject_colors)]
                        course_faculty_map[code] = {
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        }
                        color_idx += 1

                # Process all courses - both lab and non-lab
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    
                    # Skip basket courses (B1, B2, etc)
                    if not any(code.startswith(f'B{i}') for i in range(1, 10)):
                        # For same course in different sections, try to use different faculty
                        if code in course_faculty_assignments:
                            # If multiple faculty available, try to pick a different one
                            if '/' in faculty:
                                faculty_options = [f.strip() for f in faculty.split('/')] 
                                # Remove already assigned faculty
                                available_faculty = [f for f in faculty_options 
                                                     if f not in course_faculty_assignments[code]]
                                if available_faculty:
                                    faculty = available_faculty[0]
                                else:
                                    faculty = select_faculty(faculty)
                        else:
                            faculty = select_faculty(faculty)
                            course_faculty_assignments[code] = [faculty]
                    else:
                        faculty = select_faculty(faculty)
                    
                    # Calculate required slots
                    lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions = calculate_required_slots(course)
                    
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}

                    # Schedule lectures
                    for _ in range(lecture_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                            
                            # Add check for faculty-course gap
                            if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                attempts += 1
                                continue
                            
                            # Check faculty daily component limit and lecture constraints
                            if not check_faculty_daily_components(professor_schedule, faculty, day, 
                                                               department, semester, section, timetable,
                                                               code, 'LEC'):
                                attempts += 1
                                continue
                                
                            # Check availability and ensure breaks between lectures
                            slots_free = True
                            for i in range(LECTURE_DURATION):
                                current_slot = start_slot + i
                                if (current_slot in professor_schedule[faculty][day] or 
                                    timetable[day][current_slot]['type'] is not None or
                                    is_break_time(TIME_SLOTS[current_slot], semester)):
                                    slots_free = False
                                    break
                                
                                # Check for lectures before this slot
                                if current_slot > 0:
                                    if is_lecture_scheduled(timetable, day, 
                                                         max(0, current_slot - BREAK_DURATION), 
                                                         current_slot):
                                        slots_free = False
                                        break
                                
                                # Check for lectures after this slot
                                if current_slot < len(TIME_SLOTS) - 1:
                                    if is_lecture_scheduled(timetable, day,
                                                         current_slot + 1,
                                                         min(len(TIME_SLOTS), 
                                                             current_slot + BREAK_DURATION + 1)):
                                        slots_free = False
                                        break
                            
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, LECTURE_DURATION, 
                                                          rooms, batch_info, timetable, code)
                                
                                if room_id:
                                    classroom = room_id
                                    
                                    # Mark slots as used
                                    for i in range(LECTURE_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'LEC'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                                # Generate detailed reason for why this component couldn't be scheduled
                                detailed_reason = unscheduled_reason(course, department, semester, 
                                                                  professor_schedule, rooms, 'LEC', attempts)
                                
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name, 
                                                       faculty, 'LEC', 1, section, detailed_reason)
                                )

                    # Schedule tutorials
                    for _ in range(tutorial_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            
                            # Add check for faculty-course gap
                            if not check_faculty_course_gap(professor_schedule, timetable, faculty, code, day, start_slot):
                                attempts += 1
                                continue
                            
                            # Check faculty daily component limit for tutorials
                            if not check_faculty_daily_components(professor_schedule, faculty, day,
                                                               department, semester, section, timetable,
                                                               code, 'TUT'):
                                attempts += 1
                                continue
                                
                            start_slot = random.randint(0, len(TIME_SLOTS)-TUTORIAL_DURATION)
                            
                            # Check availability
                            slots_free = True
                            for i in range(TUTORIAL_DURATION):
                                if (start_slot+i in professor_schedule[faculty][day] or 
                                    timetable[day][start_slot+i]['type'] is not None or
                                    is_break_time(TIME_SLOTS[start_slot+i], semester)):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, TUTORIAL_DURATION, 
                                                          rooms, batch_info, timetable, code)
                                
                                if room_id:
                                    classroom = room_id
                                    
                                    # Mark slots as used
                                    for i in range(TUTORIAL_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'TUT'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                            # Generate detailed reason for why this tutorial couldn't be scheduled
                            detailed_reason = unscheduled_reason(course, department, semester, 
                                                              professor_schedule, rooms, 'TUT', attempts)
                            
                            unscheduled_components.add(
                                UnscheduledComponent(department, semester, code, name,
                                                   faculty, 'TUT', 1, section, detailed_reason)
                            )

                    # Schedule labs with tracking
                    if lab_sessions > 0:
                        room_type = get_required_room_type(course)
                        for _ in range(lab_sessions):
                            scheduled = False
                            attempts = 0
                            scheduling_reason = ""
                            
                            # Try each day in random order
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            
                            for day in days:
                                # Get all possible slots for this day
                                possible_slots = get_best_slots(timetable, professor_schedule, 
                                                              faculty, day, LAB_DURATION, 
                                                              semester, department)
                                
                                for start_slot in possible_slots:
                                    room_id = find_suitable_room(room_type, department, semester,
                                                               day, start_slot, LAB_DURATION,
                                                               rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id if ',' not in str(room_id) else f"{room_id.split(',')[0]}+{room_id.split(',')[1]}"
                                        
                                        # Mark slots as used
                                        for i in range(LAB_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LAB'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        break
                                
                                if scheduled:
                                    break
                                
                            if not scheduled:
                                # Generate detailed reason for why this lab component couldn't be scheduled
                                detailed_reason = unscheduled_reason(course, department, semester, 
                                                                  professor_schedule, rooms, 'LAB', attempts)
                                
                                unscheduled_components.add(
                                    UnscheduledComponent(department, semester, code, name,
                                                       faculty, 'LAB', 1, section, detailed_reason)
                                )

                # Schedule self-study sessions
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    _, _, _, self_study_sessions = calculate_required_slots(course)
                    
                    if self_study_sessions > 0:
                        if faculty not in professor_schedule:
                            professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                        
                        # Schedule each self-study session (1 hour each)
                        for _ in range(self_study_sessions):
                            scheduled = False
                            attempts = 0
                            while not scheduled and attempts < 1000:
                                day = random.randint(0, len(DAYS)-1)
                                start_slot = random.randint(0, len(TIME_SLOTS)-SELF_STUDY_DURATION)
                                
                                # Check availability
                                slots_free = True
                                for i in range(SELF_STUDY_DURATION):
                                    if (start_slot+i in professor_schedule[faculty][day] or 
                                        timetable[day][start_slot+i]['type'] is not None or
                                        is_break_time(TIME_SLOTS[start_slot+i], semester)):
                                        slots_free = False
                                        break
                                
                                if slots_free:
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                              day, start_slot, SELF_STUDY_DURATION, 
                                                              rooms, batch_info, timetable, code)
                                    
                                    if room_id:
                                        classroom = room_id
                                        
                                        # Mark slots as used
                                        for i in range(SELF_STUDY_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'SS'  # SS for Self Study
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                attempts += 1

                # Write timetable to worksheet
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                lec_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")  # Salmon
                lab_fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")  # Lawn green
                tut_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Sky blue
                ss_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")   # Gold
                break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") # Silver
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    
                    merge_ranges = []  # Track merge ranges for this row
                    
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        
                        if is_break_time(TIME_SLOTS[slot_idx], semester):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            
                            # Only create content for start of activity
                            if code:
                                # Get duration based on activity type
                                duration = {
                                    'LEC': LECTURE_DURATION,
                                    'LAB': LAB_DURATION,
                                    'TUT': TUTORIAL_DURATION,
                                    'SS': SELF_STUDY_DURATION
                                }.get(activity_type, 1)
                                
                                # Use subject-specific color
                                if code in subject_color_map:
                                    cell_fill = PatternFill(start_color=subject_color_map[code],
                                                          end_color=subject_color_map[code],
                                                          fill_type="solid")
                                else:
                                    cell_fill = {
                                        'LAB': lab_fill,
                                        'TUT': tut_fill,
                                        'SS': ss_fill,
                                        'LEC': lec_fill
                                    }.get(activity_type, lec_fill)
                                
                                if code and is_basket_course(code):
                                    basket_group = get_basket_group(code)
                                    # Get all courses from same basket in this slot
                                    basket_codes = set()  # Use set to avoid duplicates
                                    basket_details = {}
                                    
                                    # First collect all courses in this basket group
                                    for slot_id, slot_data in timetable[day_idx].items():
                                        slot_code = slot_data.get('code', '')
                                        if (slot_data.get('type') == activity_type and 
                                            get_basket_group(slot_code) == basket_group):
                                            basket_codes.add(slot_code)  # Add to set instead of list
                                            # Only store details if not already present
                                            if slot_code not in basket_details:
                                                basket_details[slot_code] = {
                                                    'faculty': slot_data['faculty'],
                                                    'room': slot_data['classroom']
                                                }
                                    
                                    if basket_codes:
                                        # Group header
                                        basket_header = f"{basket_group} Courses\n"
                                        # List of all unique course codes
                                        codes_str = ', '.join(sorted(basket_codes))
                                        # Course details with rooms (unique entries)
                                        course_details = [
                                            f"{code}: {details['faculty']} ({details['room']})"
                                            for code, details in sorted(basket_details.items())
                                            if code and details['faculty'] and details['room']
                                        ]
                                        
                                        cell_value = f"{basket_header}{codes_str}\n" + "\n".join(course_details)
                                else:
                                    cell_value = f"{code} {activity_type}\nroom no. :{classroom}\n{faculty}"
                                
                                # Create merge range
                                if duration > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + duration + 1)
                                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                                    merge_ranges.append((merge_range, cell_fill))
                        
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', indent=1)
                    
                    # Apply merges after creating all cells in the row
                    for merge_range, fill in merge_ranges:
                        ws.merge_cells(merge_range)
                        # Ensure merged cell has consistent formatting
                        merged_cell = ws[merge_range.split(':')[0]]
                        merged_cell.fill = fill
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center', indent=2)

                for col_idx in range(1, len(TIME_SLOTS)+2):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40

                # Add Self-Study Only Courses section
                current_row = len(DAYS) + 4  # Initialize current_row here, before any sections

                if self_study_courses:
                    ss_courses_for_this_section = [c for c in self_study_courses 
                                               if c['department'] == department and 
                                               c['semester'] == semester]
                    
                    if ss_courses_for_this_section:
                        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                        ws.cell(row=current_row, column=1).font = Font(bold=True)
                        current_row += 1
                        
                        headers = ['Course Code', 'Course Name', 'Faculty']
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=current_row, column=col, value=header)
                            ws.cell(row=current_row, column=col).font = Font(bold=True)
                        current_row += 1
                        
                        for course in ss_courses_for_this_section:
                            ws.cell(row=current_row, column=1, value=course['code'])
                            ws.cell(row=current_row, column=2, value=course['name'])
                            ws.cell(row=current_row, column=3, value=course['faculty'])
                            current_row += 1
                        
                        current_row += 2  # Add extra spacing after self-study courses

                # Handle unscheduled components section
                dept_unscheduled = [c for c in unscheduled_components 
                                    if c.department == department and 
                                    c.semester == semester and
                                    (c.section == section if num_sections > 1 else True)]

                if dept_unscheduled:
                    current_row += 2  # Add spacing after previous section
                    unsch_title = ws.cell(row=current_row, column=1, value="Unscheduled Components")
                    unsch_title.font = Font(bold=True, size=12, color="FF0000")
                    current_row += 2

                    headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Reason']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        # Set column widths for better readability
                        ws.column_dimensions[get_column_letter(col)].width = 20
                    current_row += 1

                    for comp in dept_unscheduled:
                        cells = [
                            (comp.code, None),
                            (comp.name, None),
                            (comp.faculty, None),
                            (comp.component_type, None),
                            (comp.sessions, None),
                            (comp.reason or "Could not find suitable slot", None)
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        current_row += 1
                    
                    current_row += 2  # Add spacing before legend

                # Improved legend formatting
                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2

                # Wider columns for legend
                ws.column_dimensions['A'].width = 20  # Subject Code
                ws.column_dimensions['B'].width = 10  # Color (moved next to code)
                ws.column_dimensions['C'].width = 40  # Subject Name
                ws.column_dimensions['D'].width = 30  # Faculty
                ws.column_dimensions['E'].width = 15  # LTPS

                # Add legend headers with better formatting
                legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Add padding to header cells
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                current_row += 1

                # Add subject entries with improved spacing and color next to code
                for code, color in subject_color_map.items():
                    if code in course_faculty_map:
                        # Add more spacing between rows
                        ws.row_dimensions[current_row].height = 30
                        
                        # Get LTPS values for this course
                        ltps_value = ""
                        for _, course_row in courses.iterrows():
                            if str(course_row['Course Code']) == code:
                                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                                ltps_value = f"{l}-{t}-{p}-{s}"
                                break
                        
                        # Create cells with padding and color next to code
                        cells = [
                            (code, None),
                            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
                            (course_faculty_map[code]['name'], None),
                            (course_faculty_map[code]['faculty'], None),
                            (ltps_value, None)
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            # Add padding with increased wrap_text and adjusted alignment
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)
                        
                        current_row += 1

    # Format the overview sheet
    for col in range(1, 4):
        overview_sheet.column_dimensions[get_column_letter(col)].width = 20
    
    for row in overview_sheet.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.font = Font(bold=True)
    
    # Apply formatting to the overview table headers
    for cell in overview_sheet[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Apply borders to the overview data
    for row in overview_sheet.iter_rows(min_row=5, max_row=row_index-1):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))

    # Save the workbook
    wb.save("timetable_all_departments.xlsx")
    print("Combined timetable for all departments and semesters saved as timetable_all_departments.xlsx")
    
    return ["timetable_all_departments.xlsx"]

def check_unscheduled_courses():
    """Check and print courses that are not scheduled according to their L-T-P-S requirements"""
    try:
        df = pd.read_csv('combined.csv')
        
        # Check if timetable file exists first
        if not os.path.exists('timetable_all_departments.xlsx'):
            print("Warning: timetable_all_departments.xlsx not found. Run generate_all_timetables() first.")
            return

        # Load the generated timetable
        wb = pd.ExcelFile('timetable_all_departments.xlsx')
        
        # Dictionary to track scheduled hours for each course
        scheduled_hours = defaultdict(lambda: {'L': 0, 'T': 0, 'P': 0, 'S': 0})
        
        # Process each sheet in the timetable workbook to extract all scheduled courses
        found_courses = []
        
        print("\nExamining timetable for scheduled courses...")
        
        for sheet_name in wb.sheet_names:
            timetable_df = pd.read_excel(wb, sheet_name=sheet_name)
            
            # Skip if this is not a valid timetable sheet (should have at least 'Day' column)
            if 'Day' not in timetable_df.columns:
                continue
            
            print(f"Processing sheet: {sheet_name}")
            
            # Process each row (day) in the timetable
            for _, row in timetable_df.iterrows():
                day = row['Day']
                if day not in DAYS:  # Skip header or legend rows
                    continue
                
                # Process each time slot in the day
                for col in timetable_df.columns[1:]:  # Skip the 'Day' column
                    cell_value = row[col]
                    
                    # Skip empty cells or break cells
                    if pd.isna(cell_value) or cell_value == '' or cell_value == 'BREAK':
                        continue
                    
                    # Process cell content to extract course information
                    # Format example: "CS101 LEC room no. :A101"
                    if isinstance(cell_value, str) and 'room no.' in cell_value:
                        # Split the cell value to extract components
                        parts = cell_value.split('room no.')
                        course_info = parts[0].strip()
                        
                        # Extract course code and class type
                        if ' ' in course_info:
                            course_parts = course_info.split()
                            if len(course_parts) >= 2:
                                course_code = course_parts[0].strip()
                                class_type = course_parts[1].strip()
                                
                                # Debug output for problematic courses
                                if 'HS204' in course_code or 'HS153' in course_code:
                                    print(f"Found in timetable: {course_code}, Type: {class_type}, Day: {day}, Slot: {col}")
                                
                                found_courses.append(course_code)  # Track all courses found in timetable
                                
                                # Update scheduled hours based on class type
                                if class_type == 'LEC':
                                    scheduled_hours[course_code]['L'] += 1.5  # Lecture is 1.5 hours
                                    print(f"Added 1.5 lecture hours for {course_code}")
                                elif class_type == 'TUT':
                                    scheduled_hours[course_code]['T'] += 1    # Tutorial is 1 hour
                                    print(f"Added 1 tutorial hour for {course_code}")
                                elif class_type == 'LAB':
                                    scheduled_hours[course_code]['P'] += 2    # Lab is 2 hours
                                    print(f"Added 2 practical hours for {course_code}")
        
        # Create mappings for course codes with aliases or variations
        all_found_courses = set(found_courses)
        course_primary_codes = {}  # Maps each course code variant to its primary code
        
        # Build a mapping of all course code variations
        for _, course in df.iterrows():
            original_code = str(course['Course Code']).strip()
            
            # For courses with slashes (e.g., "HS204 / HS153")
            if '/' in original_code:
                variants = [c.strip() for c in original_code.split('/')]
                primary = variants[0]
                
                # Map all variants to the primary code
                for variant in variants:
                    course_primary_codes[variant] = primary
                
                # Also map the complete original code
                course_primary_codes[original_code] = primary
                
                # Debug output for specific courses
                if 'HS204' in original_code or 'HS153' in original_code:
                    print(f"Mapping variants for {original_code}: {variants}")
            
            # For courses with parentheses (e.g., "B1(ASD151/HS151/New/New/New/New)")
            elif '(' in original_code and ')' in original_code:
                base = original_code.split('(')[0].strip()
                inner_part = original_code.split('(')[1].split(')')[0]
                
                if '/' in inner_part:
                    inner_variants = [c.strip() for c in inner_part.split('/')]
                    for variant in inner_variants:
                        if variant.lower() != 'new':
                            # Create combined codes like "B1_ASD151"
                            combined = f"{base}_{variant}"
                            course_primary_codes[combined] = original_code
                            course_primary_codes[variant] = original_code
                
                # Map the original code to itself
                course_primary_codes[original_code] = original_code
            
            # Regular codes just map to themselves
            else:
                course_primary_codes[original_code] = original_code
        
        # Print courses found in timetable for debugging
        print("\nUnique courses found in timetable:", len(all_found_courses))
        print("First 10 courses found:", list(all_found_courses)[:10])
        
        # Merge scheduled hours for course variants
        merged_hours = defaultdict(lambda: {'L': 0, 'T': 0, 'P': 0, 'S': 0})
        
        for code, hours in scheduled_hours.items():
            # Find the primary code (or use the code itself if not a variant)
            primary_code = course_primary_codes.get(code, code)
            
            # Debug output for problematic courses
            if 'HS204' in code or 'HS153' in code or 'HS204' in primary_code or 'HS153' in primary_code:
                print(f"Merging hours for {code} -> {primary_code}: {hours}")
            
            # Accumulate hours by primary code
            merged_hours[primary_code]['L'] += hours['L']
            merged_hours[primary_code]['T'] += hours['T']
            merged_hours[primary_code]['P'] += hours['P']
            merged_hours[primary_code]['S'] += hours['S']
        
        # Compare with required hours to find unscheduled courses
        unscheduled_courses = []
        
        for _, course in df.iterrows():
            original_code = str(course['Course Code']).strip()
            name = str(course['Course Name'])
            faculty = str(course['Faculty'])
            department = str(course['Department'])
            semester = str(course['Semester'])
            
            # Get the primary code for looking up scheduled hours
            primary_code = course_primary_codes.get(original_code, original_code)
            
            # Extract LTPS requirements
            required_l = int(course['L']) if pd.notna(course['L']) else 0
            required_t = int(course['T']) if pd.notna(course['T']) else 0
            required_p = int(course['P']) if pd.notna(course['P']) else 0
            required_s = int(course['S']) if pd.notna(course['S']) and 'S' in course else 0
            
            # Get scheduled hours for this course using the primary code
            scheduled_l = merged_hours[primary_code]['L']
            scheduled_t = merged_hours[primary_code]['T']
            scheduled_p = merged_hours[primary_code]['P']
            scheduled_s = merged_hours[primary_code]['S']
            
            # Debug output for specific courses
            if 'HS204' in original_code or 'HS153' in original_code:
                print(f"\nCourse: {original_code} (Primary: {primary_code})")
                print(f"  Required L-T-P-S: {required_l}-{required_t}-{required_p}-{required_s}")
                print(f"  Merged Scheduled L-T-P-S: {scheduled_l}-{scheduled_t}-{scheduled_p}-{scheduled_s}")
            
            # Use a tolerance for floating-point comparisons
            tolerance = 0.01
            
            # Check if any component is under-scheduled
            missing_l = max(0, required_l - scheduled_l)
            missing_t = max(0, required_t - scheduled_t)
            missing_p = max(0, required_p - scheduled_p)
            missing_s = max(0, required_s - scheduled_s)
            
            if (missing_l > tolerance or missing_t > tolerance or 
                missing_p > tolerance or missing_s > tolerance):
                
                # Determine possible reasons for unscheduled components
                reasons = []
                
                # Check if any variant of this course was found in the timetable
                variants_found = False
                found_variants = []
                
                # Check all variants
                if primary_code in course_primary_codes.values():
                    for code, primary in course_primary_codes.items():
                        if primary == primary_code and code in all_found_courses:
                            variants_found = True
                            found_variants.append(code)
                else:
                    # Check the primary code itself
                    if primary_code in all_found_courses:
                        variants_found = True
                        found_variants.append(primary_code)
                
                if variants_found:
                    reasons.append(f"Course found in timetable as {', '.join(found_variants)} but not all required hours are scheduled")
                else:
                    reasons.append("Course not found in any timetable")
                
                # Check for faculty conflicts
                faculty_courses = df[df['Faculty'] == faculty]['Course Code'].tolist()
                if len(faculty_courses) > 1:
                    reasons.append("Faculty teaching multiple courses may have scheduling constraints")
                
                # Check for room availability
                section_key = (department, semester)
                if len(lecture_rooms) < 1 and required_l > 0:
                    reasons.append("Insufficient lecture rooms available")
                if len(computer_lab_rooms) < 1 and required_p > 0:
                    reasons.append("Insufficient lab rooms available")
                
                # Check for semester course load
                semester_courses = df[(df['Department'] == department) & (df['Semester'] == semester)].shape[0]
                if semester_courses > 6:
                    reasons.append("High number of courses in same semester may cause conflicts")
                
                # Add course to unscheduled list
                unscheduled_courses.append({
                    'Code': original_code,
                    'Name': name,
                    'Faculty': faculty,
                    'Department': department,
                    'Semester': semester,
                    'Required L-T-P-S': f"{required_l}-{required_t}-{required_p}-{required_s}",
                    'Scheduled L-T-P-S': f"{scheduled_l}-{scheduled_t}-{scheduled_p}-{scheduled_s}",
                    'Missing L': round(missing_l, 2),
                    'Missing T': round(missing_t, 2),
                    'Missing P': round(missing_p, 2),
                    'Missing S': round(missing_s, 2),
                    'Variant Found': variants_found,
                    'Found As': ', '.join(found_variants) if found_variants else "Not found",
                    'Reasons': "; ".join(reasons)
                })
        
        # Print results
        if unscheduled_courses:
            print("\n=== COURSES WITH UNSCHEDULED HOURS ===")
            print(f"Found {len(unscheduled_courses)} courses with scheduling issues:\n")
            
            for course in unscheduled_courses:
                print(f"Course: {course['Code']} - {course['Name']}")
                print(f"  Department: {course['Department']}, Semester: {course['Semester']}")
                print(f"  Faculty: {course['Faculty']}")
                print(f"  Required L-T-P-S: {course['Required L-T-P-S']}")
                print(f"  Scheduled L-T-P-S: {course['Scheduled L-T-P-S']}")
                
                missing = []
                if course['Missing L'] > 0:
                    missing.append(f"{course['Missing L']} lecture hours")
                if course['Missing T'] > 0:
                    missing.append(f"{course['Missing T']} tutorial hours")
                if course['Missing P'] > 0:
                    missing.append(f"{course['Missing P']} practical hours")
                if course['Missing S'] > 0:
                    missing.append(f"{course['Missing S']} self-study hours")
                    
                print(f"  Missing: {', '.join(missing)}")
                print(f"  Found in Timetable: {'Yes' if course['Variant Found'] else 'No'}")
                if course['Variant Found']:
                    print(f"  Found as: {course['Found As']}")
                print(f"  Possible Reasons: {course['Reasons']}\n")
            
            # Create Excel file with unscheduled courses
            unscheduled_df = pd.DataFrame(unscheduled_courses)
            unscheduled_df.to_excel('unscheduled_courses.xlsx', index=False)
            print("Details saved to 'unscheduled_courses.xlsx'")
        else:
            print("\n=== ALL COURSES FULLY SCHEDULED ===")
            print("All courses have been scheduled according to their L-T-P-S requirements.")
            
    except Exception as e:
        print(f"Error checking unscheduled courses: {e}")
        import traceback
        traceback.print_exc()

def generate_faculty_timetables():
    """Generate timetables for all faculty members in a single Excel file"""
    try:
        # Load the generated timetable
        wb = pd.ExcelFile('timetable_all_departments.xlsx')
        
        # Dictionary to track faculty schedules
        faculty_schedules = {}
        
        print("Processing timetable sheets to extract faculty schedules...")
        
        # Load courses data for reference
        try:
            courses_data = pd.read_csv('combined.csv')
        except Exception as e:
            print(f"Warning: Could not load combined.csv for reference: {e}")
            courses_data = None
        
        # Process each sheet in the timetable workbook
        for sheet_name in wb.sheet_names:
            timetable_df = pd.read_excel(wb, sheet_name=sheet_name)
            
            # Skip if this is not a valid timetable sheet (should have at least 'Day' column)
            if 'Day' not in timetable_df.columns:
                continue
                
            print(f"Processing sheet: {sheet_name}")
            
            # Extract department and semester from sheet name
            dept_sem = sheet_name
                
            # Process each row (day) in the timetable
            for _, row in timetable_df.iterrows():
                day = row['Day']
                if day not in DAYS:  # Skip header or legend rows
                    continue
                    
                # Process each time slot in the day
                for col in timetable_df.columns[1:]:  # Skip the 'Day' column
                    cell_value = row[col]
                    
                    # Skip empty cells or break cells
                    if pd.isna(cell_value) or cell_value == '' or cell_value == 'BREAK':
                        continue
                    
                    # Debug print for cell content inspection
                    # print(f"Cell content: {cell_value}")
                    
                    # First pattern: "COURSE_CODE CLASS_TYPE\nroom no. :ROOM\nFACULTY_NAME"
                    if isinstance(cell_value, str) and "room no." in cell_value:
                        try:
                            lines = cell_value.strip().split('\n')
                            
                            # First line contains course code and class type
                            first_line_parts = lines[0].split()
                            if len(first_line_parts) >= 2:
                                course_code = first_line_parts[0]
                                class_type = first_line_parts[1]
                                
                                # Room information is in the second line
                                room_info = lines[1].replace('room no. :', '').strip() if len(lines) > 1 else "Unknown"
                                
                                # Faculty is in the third line
                                faculty_name = lines[2].strip() if len(lines) > 2 else ""
                                
                                # If faculty name is empty or too short, try to get it from courses_data
                                if not faculty_name or len(faculty_name) < 2:
                                    if courses_data is not None:
                                        course_row = courses_data[courses_data['Course Code'] == course_code]
                                        if not course_row.empty:
                                            faculty_name = str(course_row['Faculty'].iloc[0])
                                
                                # Get course name from courses_data
                                course_name = ""
                                if courses_data is not None:
                                    course_row = courses_data[courses_data['Course Code'] == course_code]
                                    if not course_row.empty:
                                        course_name = str(course_row['Course Name'].iloc[0])
                                
                                # Process each faculty
                                if faculty_name:
                                    faculty_list = extract_faculty_names(faculty_name)
                                    
                                    for faculty in faculty_list:
                                        if faculty.strip():  # Only process non-empty faculty names
                                            # Initialize faculty entry if not already exists
                                            if faculty not in faculty_schedules:
                                                faculty_schedules[faculty] = {d: {} for d in DAYS}
                                            
                                            # Add this class to faculty schedule
                                            time_slot_str = col
                                            faculty_schedules[faculty][day][time_slot_str] = {
                                                'Course Code': course_code,
                                                'Course Name': course_name,
                                                'Class Type': class_type,
                                                'Room': room_info,
                                                'Department-Semester': dept_sem
                                            }
                        except Exception as e:
                            print(f"Error processing cell: {cell_value}")
                            print(f"Error details: {e}")
                            traceback.print_exc()
                    
                    # Second pattern: Handle basket course format
                    elif isinstance(cell_value, str) and "Courses" in cell_value:
                        try:
                            lines = cell_value.strip().split('\n')
                            
                            # Skip the header line
                            basket_courses = []
                            basket_details = []
                            
                            # Process each line to extract course details
                            for line in lines[1:]:
                                if ':' in line:  # This is a detail line with faculty and room
                                    parts = line.split(':')
                                    if len(parts) >= 2:
                                        code = parts[0].strip()
                                        details = parts[1].strip()
                                        
                                        # Extract faculty and room
                                        if '(' in details and ')' in details:
                                            faculty_part = details.split('(')[0].strip()
                                            room_part = details.split('(')[1].split(')')[0].strip()
                                            
                                            basket_details.append({
                                                'code': code,
                                                'faculty': faculty_part,
                                                'room': room_part
                                            })
                                elif ',' in line and not any(x in line for x in ['Courses', 'room no.']):
                                    # This might be the course codes line
                                    basket_courses = [code.strip() for code in line.split(',')]
                            
                            # Process each basket course
                            for detail in basket_details:
                                code = detail['code']
                                faculty_name = detail['faculty']
                                room_info = detail['room']
                                
                                # Get course name from courses_data
                                course_name = ""
                                if courses_data is not None:
                                    course_row = courses_data[courses_data['Course Code'] == code]
                                    if not course_row.empty:
                                        course_name = str(course_row['Course Name'].iloc[0])
                                
                                # Determine class type from the original cell
                                class_type = "LEC"  # Default
                                if "LAB" in cell_value:
                                    class_type = "LAB"
                                elif "TUT" in cell_value:
                                    class_type = "TUT"
                                
                                # Process faculty names
                                faculty_list = extract_faculty_names(faculty_name)
                                
                                for faculty in faculty_list:
                                    if faculty.strip():  # Only process non-empty faculty names
                                        # Initialize faculty entry if not already exists
                                        if faculty not in faculty_schedules:
                                            faculty_schedules[faculty] = {d: {} for d in DAYS}
                                        
                                        # Add this class to faculty schedule
                                        time_slot_str = col
                                        faculty_schedules[faculty][day][time_slot_str] = {
                                            'Course Code': code,
                                            'Course Name': course_name,
                                            'Class Type': class_type,
                                            'Room': room_info,
                                            'Department-Semester': dept_sem
                                        }
                        except Exception as e:
                            print(f"Error processing basket cell: {cell_value}")
                            print(f"Error details: {e}")
        
        # Create a single Excel workbook for all faculty timetables
        print(f"Creating a consolidated Excel file with {len(faculty_schedules)} faculty timetables...")
        faculty_wb = Workbook()
        
        # Remove the default sheet
        if "Sheet" in faculty_wb.sheetnames:
            faculty_wb.remove(faculty_wb["Sheet"])
        
        # Create an overview/index sheet
        overview = faculty_wb.create_sheet("Overview", 0)
        overview.column_dimensions['A'].width = 40
        overview.column_dimensions['B'].width = 15
        
        # Add title and headers
        overview.append(["Faculty Timetable - All Faculty"])
        overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        overview.append([])
        overview.append(["Faculty Name", "Total Classes"])
        
        # Apply formatting to headers
        for row in range(1, 5):
            for cell in overview[row]:
                cell.font = Font(bold=True)
        
        # Style the header row
        for cell in overview[4]:
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add faculty list to overview with hyperlinks to their sheets
        row_idx = 5
        for faculty in sorted(faculty_schedules.keys()):
            total_classes = len(sum([list(slots.keys()) for slots in faculty_schedules[faculty].values()], []))
            
            # Add to overview sheet
            overview.cell(row=row_idx, column=1, value=faculty)
            overview.cell(row=row_idx, column=2, value=total_classes)
            
            # Add hyperlink to the faculty's sheet
            safe_name = sanitize_sheet_name(faculty)
            overview.cell(row=row_idx, column=1).hyperlink = f"#{safe_name}!A1"
            overview.cell(row=row_idx, column=1).style = "Hyperlink"
            
            # Add border to cells
            for col in range(1, 3):
                overview.cell(row=row_idx, column=col).border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            row_idx += 1
        
        # Generate worksheets for each faculty
        for i, faculty in enumerate(sorted(faculty_schedules.keys())):
            # Create a sanitized sheet name (Excel has a 31 character limit for sheet names)
            sheet_name = sanitize_sheet_name(faculty)
            
            # Create a new worksheet for this faculty
            ws = faculty_wb.create_sheet(title=sheet_name)
            
            # Generate the faculty's timetable in this worksheet
            create_faculty_worksheet(ws, faculty, faculty_schedules[faculty])
            
            if i % 10 == 0:  # Print progress every 10 faculty
                print(f"Generated {i+1}/{len(faculty_schedules)} faculty worksheets")
        
        # Save the workbook
        faculty_wb.save("all_faculty_timetables.xlsx")
        print(f"All {len(faculty_schedules)} faculty timetables saved in 'all_faculty_timetables.xlsx'")
        
    except Exception as e:
        print(f"Error generating faculty timetables: {e}")
        traceback.print_exc()

def sanitize_sheet_name(name):
    """Create a valid Excel sheet name from a faculty name"""
    # Replace invalid sheet name characters
    invalid_chars = ['/', '\\', '?', '*', ':', '[', ']', "'", '"']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    
    # Excel has 31 character limit for sheet names
    if len(sanitized) > 31:
        sanitized = sanitized[:28] + "..."
    
    return sanitized

def create_faculty_worksheet(ws, faculty, schedule):
    """Create a worksheet for a faculty member's timetable"""
    # Add faculty name as title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"Schedule for: {faculty}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Create header (starting from row 2)
    header = ['Day', 'Time Slot', 'Course Code', 'Course Name', 'Class Type', 'Room', 'Department-Semester']
    ws.append(header)
    
    # Apply header formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add each scheduled class for this faculty
    row_idx = 3
    for day in DAYS:
        # Sort time slots chronologically
        time_slots = sorted(schedule[day].keys())
        
        if not time_slots:  # No classes on this day
            ws.append([day, "No classes scheduled", "", "", "", "", ""])
            row_idx += 1
            continue
            
        for time_slot in time_slots:
            class_info = schedule[day][time_slot]
            
            # Add this class to the worksheet
            ws.append([
                day,
                time_slot,
                class_info['Course Code'],
                class_info['Course Name'],
                class_info['Class Type'],
                class_info['Room'],
                class_info['Department-Semester']
            ])
            
            # Apply formatting
            for cell in ws[row_idx]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Highlight each class type differently
            class_type = class_info['Class Type']
            if class_type == 'LEC':
                fill_color = "B8CCE4"  # Light blue
            elif class_type == 'TUT':
                fill_color = "E4B8CC"  # Pink
            elif class_type == 'LAB':
                fill_color = "CCE4B8"  # Light green
            else:
                fill_color = "F2F2F2"  # Light gray
                
            # Apply fill color
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            row_idx += 1
    
    # Set column widths
    column_widths = {
        'A': 15,  # Day
        'B': 20,  # Time Slot
        'C': 15,  # Course Code
        'D': 40,  # Course Name
        'E': 12,  # Class Type
        'F': 25,  # Room
        'G': 30,  # Department-Semester
    }
    
    # Apply the predefined widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

# Enhance extract_faculty_names function to handle more patterns
def extract_faculty_names(faculty_string):
    """Extract individual faculty names from a combined string"""
    if not faculty_string or pd.isna(faculty_string):
        return []
        
    faculty_string = str(faculty_string).strip()
    if faculty_string.lower() in ['nan', 'none', '']:
        return []
        
    faculty_names = []
    
    # Split by common separators and handle various formats
    if '&' in faculty_string:
        parts = faculty_string.split('&')
        for part in parts:
            faculty_names.append(part.strip())
    elif ' and ' in faculty_string.lower():
        # Split by "and" but be careful with words that might contain "and"
        parts = faculty_string.lower().split(' and ')
        for i, part in enumerate(parts):
            # Reconstruct the original case
            start_idx = faculty_string.lower().find(part)
            if start_idx >= 0:
                end_idx = start_idx + len(part)
                faculty_names.append(faculty_string[start_idx:end_idx].strip())
    elif ',' in faculty_string and faculty_string.count(',') > 1:
        # Multiple commas likely indicate a list of names
        parts = faculty_string.split(',')
        for part in parts:
            if part.strip():  # Skip empty parts
                faculty_names.append(part.strip())
    elif '/' in faculty_string:
        parts = faculty_string.split('/')
        for part in parts:
            faculty_names.append(part.strip())
    elif ';' in faculty_string:
        parts = faculty_string.split(';')
        for part in parts:
            faculty_names.append(part.strip())
    else:
        faculty_names.append(faculty_string)  # Single faculty
    
    # Remove any empty strings and normalize
    return [name.strip() for name in faculty_names if name.strip()]

def generate_individual_faculty_timetable(faculty, schedule):
    """Generate a timetable for a single faculty member"""
    # Sanitize faculty name for file name
    filename = sanitize_filename(faculty)
    file_path = os.path.join('faculty_timetables', f"timetable_{filename}.xlsx")
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Add faculty name as title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"Schedule for: {faculty}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Create header (starting from row 2)
    header = ['Day', 'Time Slot', 'Course Code', 'Course Name', 'Class Type', 'Room', 'Department-Semester']
    ws.append(header)
    
    # Apply header formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add each scheduled class for this faculty
    row_idx = 3
    for day in DAYS:
        # Sort time slots chronologically
        time_slots = sorted(schedule[day].keys())
        
        if not time_slots:  # No classes on this day
            ws.append([day, "No classes scheduled", "", "", "", "", ""])
            row_idx += 1
            continue
            
        for time_slot in time_slots:
            class_info = schedule[day][time_slot]
            
            # Add this class to the worksheet
            ws.append([
                day,
                time_slot,
                class_info['Course Code'],
                class_info['Course Name'],
                class_info['Class Type'],
                class_info['Room'],
                class_info['Department-Semester']
            ])
            
            # Apply formatting
            for cell in ws[row_idx]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Highlight each class type differently
            class_type = class_info['Class Type']
            if class_type == 'LEC':
                fill_color = "B8CCE4"  # Light blue
            elif class_type == 'TUT':
                fill_color = "E4B8CC"  # Pink
            elif class_type == 'LAB':
                fill_color = "CCE4B8"  # Light green
            else:
                fill_color = "F2F2F2"  # Light gray
                
            # Apply fill color
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            row_idx += 1
    
    # Set column widths
    column_widths = {
        'A': 15,  # Day
        'B': 20,  # Time Slot
        'C': 15,  # Course Code
        'D': 40,  # Course Name
        'E': 12,  # Class Type
        'F': 25,  # Room
        'G': 30,  # Department-Semester
    }
    
    # Apply the predefined widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Save the workbook
    wb.save(file_path)

def sanitize_filename(name):
    """Create a valid file name from a faculty name"""
    # Replace invalid filename characters
    invalid_chars = ['/', '\\', '?', '*', ':', '[', ']', "'", '"', '<', '>', '|', ' ']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    
    # Remove consecutive underscores
    while '__' in sanitized:
        sanitized = sanitized.replace('__', '_')
    
    # Limit length to avoid overly long filenames
    if len(sanitized) > 50:
        sanitized = sanitized[:50]
        
    # Remove trailing underscores
    sanitized = sanitized.rstrip('_')
    
    return sanitized

# Initialize global variables
TIME_SLOTS = []
lunch_breaks = {}  # Global lunch breaks dictionary

def load_config():
    """Load duration constants from config file"""
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
            return config['duration_constants']
    except:
        # Return defaults if config file not found
        return {
            'hour_slots': 2,
            'lecture_duration': 3,
            'lab_duration': 4,
            'tutorial_duration': 2,
            'self_study_duration': 2, 
            'break_duration': 1
        }

def initialize_time_slots():
    """Initialize time slots for scheduling"""
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()

def calculate_lunch_breaks(semesters):
    """Dynamically calculate staggered lunch breaks for semesters"""
    global lunch_breaks
    lunch_breaks = {}  # Reset global lunch_breaks
    total_semesters = len(semesters)
    
    if total_semesters == 0:
        return lunch_breaks
        
    # Calculate time between breaks to distribute them evenly
    total_window_minutes = (
        LUNCH_WINDOW_END.hour * 60 + LUNCH_WINDOW_END.minute -
        LUNCH_WINDOW_START.hour * 60 - LUNCH_WINDOW_START.minute
    )
    stagger_interval = (total_window_minutes - LUNCH_DURATION) / (total_semesters - 1) if total_semesters > 1 else 0
    
    # Sort semesters to ensure consistent assignment
    sorted_semesters = sorted(semesters)
    
    for i, semester in enumerate(sorted_semesters):
        start_minutes = (LUNCH_WINDOW_START.hour * 60 + LUNCH_WINDOW_START.minute + 
                        int(i * stagger_interval))
        start_hour = start_minutes // 60
        start_min = start_minutes % 60
        
        end_minutes = start_minutes + LUNCH_DURATION
        end_hour = end_minutes // 60
        end_min = end_minutes % 60
        
        lunch_breaks[semester] = (
            time(start_hour, start_min),
            time(end_hour, end_min)
        )
    
    return lunch_breaks

def is_break_time(slot, semester=None):
    """Check if a time slot falls within break times"""
    global lunch_breaks
    start, end = slot
    
    # Morning break: 10:30-11:00
    morning_break = (time(10, 30) <= start < time(11, 0))
    
    # Staggered lunch breaks based on semester
    lunch_break = False
    if semester:
        base_sem = int(str(semester)[0])  # Get base semester number (e.g., 4 from 4A)
        if base_sem in lunch_breaks:
            lunch_start, lunch_end = lunch_breaks[base_sem]
            lunch_break = (lunch_start <= start < lunch_end)
    else:
        # For general checks without semester info, block all lunch periods
        lunch_break = any(lunch_start <= start < lunch_end 
                         for lunch_start, lunch_end in lunch_breaks.values())
    
    return morning_break or lunch_break

def load_rooms():
    """Load room information from CSV file"""
    rooms = {}
    try:
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(DAYS))}
                }
    except FileNotFoundError:
        print("Warning: rooms.csv not found, using default room allocation")
        return None
    return rooms

def load_batch_data():
    """Load batch information and calculate sections automatically using total_students from combined.csv"""
    batch_info = {}
    
    # Load batch sizes directly from combined.csv
    try:
        df = pd.read_csv('combined.csv')
        
        # Group by Department and Semester to get total students
        grouped = df.groupby(['Department', 'Semester'])
        
        for (dept, sem), group in grouped:
            # Check if total_students column exists and has values
            if 'total_students' in group.columns:
                # Get the max number of students for this department/semester
                # Filter out non-numeric values and convert to integers
                valid_students = []
                for val in group['total_students']:
                    try:
                        if pd.notna(val) and str(val).isdigit():
                            valid_students.append(int(val))
                    except (ValueError, TypeError):
                        continue
                
                if valid_students:
                    total_students = max(valid_students)
                    
                    # Default max batch size of 70
                    max_batch_size = 70
                    
                    # Calculate number of sections needed
                    num_sections = (total_students + max_batch_size - 1) // max_batch_size
                    section_size = (total_students + num_sections - 1) // num_sections

                    batch_info[(dept, sem)] = {
                        'total': total_students,
                        'num_sections': num_sections,
                        'section_size': section_size
                    }
                
        # Process basket/elective courses individually
        basket_courses = df[df['Course Code'].astype(str).str.contains('^B[0-9]')]
        
        # Process each basket course
        for _, course in basket_courses.iterrows():
            code = str(course['Course Code'])
            # Use total_students column if available and is a valid number
            if 'total_students' in df.columns:
                try:
                    val = course['total_students']
                    if pd.notna(val) and str(val).isdigit():
                        total_students = int(val)
                    else:
                        # Default to 35 students for basket courses if not a valid number
                        total_students = 35
                except (ValueError, TypeError):
                    total_students = 35
            else:
                # Default to 35 students for basket courses if column not available
                total_students = 35
                
            batch_info[('ELECTIVE', code)] = {
                'total': total_students,
                'num_sections': 1,  # Electives are typically single section
                'section_size': total_students
            }
            
    except FileNotFoundError:
        print("Warning: combined.csv not found, using default batch sizes")
    except Exception as e:
        print(f"Warning: Error processing batch data from combined.csv: {e}")
        
    return batch_info

def is_basket_course(code):
    """Check if course is part of a basket based on code prefix"""
    return code.startswith('B') and '-' in code

def get_basket_group(code):
    """Get the basket group (B1, B2 etc) from course code"""
    if is_basket_course(code):
        return code.split('-')[0]
    return None

def get_basket_group_slots(timetable, day, basket_group):
    """Find existing slots with courses from same basket group"""
    basket_slots = []
    for slot_idx, slot in timetable[day].items():
        code = slot.get('code', '')
        if code and get_basket_group(code) == basket_group:
            basket_slots.append(slot_idx)
    return basket_slots

def find_adjacent_lab_room(room_id, rooms):
    """Find an adjacent lab room based on room numbering"""
    if not room_id:
        return None
    
    # Get room number and extract base info
    current_num = int(''.join(filter(str.isdigit, rooms[room_id]['roomNumber'])))
    current_floor = current_num // 100
    
    # Look for adjacent room with same type
    for rid, room in rooms.items():
        if rid != room_id and room['type'] == rooms[room_id]['type']:
            room_num = int(''.join(filter(str.isdigit, room['roomNumber'])))
            # Check if on same floor and adjacent number
            if room_num // 100 == current_floor and abs(room_num - current_num) == 1:
                return rid
    return None

def try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids):
    """Helper function to try allocating rooms of a certain type"""
    for room_id, room in rooms.items():
        if room_id in used_room_ids or room['type'].upper() == 'LIBRARY':
            continue
            
        # For lectures and tutorials, only use lecture rooms and seater rooms
        if course_type in ['LEC', 'TUT', 'SS']:
            if not ('LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()):
                continue
        # For labs, match lab type exactly
        elif course_type == 'COMPUTER_LAB' and room['type'].upper() != 'COMPUTER_LAB':
            continue
        elif course_type == 'HARDWARE_LAB' and room['type'].upper() != 'HARDWARE_LAB':
            continue
            
        # Check capacity except for labs which can be split into batches
        if course_type not in ['COMPUTER_LAB', 'HARDWARE_LAB'] and room['capacity'] < required_capacity:
            continue

        # Check availability
        slots_free = True
        for i in range(duration):
            if start_slot + i in room['schedule'][day]:
                slots_free = False
                break
                
        if slots_free:
            for i in range(duration):
                room['schedule'][day].add(start_slot + i)
            return room_id
                
    return None

def get_required_room_type(course):
    """Determine required room type based on course attributes"""
    if pd.notna(course['P']) and course['P'] > 0:
        course_code = str(course['Course Code']).upper()
        # For CS courses, use computer labs
        if 'CS' in course_code or 'DS' in course_code:
            return 'COMPUTER_LAB'
        # For EC courses, use hardware labs
        elif 'EC' in course_code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'  # Default to computer lab if unspecified
    else:
        # For lectures, tutorials, and self-study
        return 'LECTURE_ROOM'

def find_suitable_room(course_type, department, semester, day, start_slot, duration, rooms, batch_info, timetable, course_code="", used_rooms=None):
    """Find suitable room(s) considering student numbers and avoiding room conflicts"""
    if not rooms:
        return "DEFAULT_ROOM"
    
    required_capacity = 60  # Default fallback
    is_basket = is_basket_course(course_code)
    total_students = None
    
    try:
        # Get total_students from combined.csv for the course
        df = pd.read_csv('combined.csv')
        
        if course_code and not is_basket:
            # For regular courses, get total_students from the course row
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns:
                total_students = int(course_row['total_students'].iloc[0])
        elif is_basket:
            # For basket courses, get total_students from the course row
            course_row = df[df['Course Code'] == course_code]
            if not course_row.empty and 'total_students' in course_row.columns:
                total_students = int(course_row['total_students'].iloc[0])
            else:
                # Fallback to batch_info if course not found directly
                elective_info = batch_info.get(('ELECTIVE', course_code))
                if elective_info:
                    total_students = elective_info['section_size']
        else:
            # If no course code (should not happen), fallback to dept info
            dept_info = batch_info.get((department, semester))
            if dept_info:
                total_students = dept_info['section_size']
    except Exception as e:
        print(f"Warning: Error getting total_students from combined.csv: {e}")
    
    # If we have total_students, use it, otherwise fallback to batch_info
    if total_students:
        required_capacity = total_students
    elif batch_info:
        # Fallbacks using batch_info
        if is_basket:
            elective_info = batch_info.get(('ELECTIVE', course_code))
            if elective_info:
                required_capacity = elective_info['section_size']
        else:
            dept_info = batch_info.get((department, semester))
            if dept_info:
                required_capacity = dept_info['section_size']

    used_room_ids = set() if used_rooms is None else used_rooms

    # Special handling for large classes based on total_students
    if course_type in ['LEC', 'TUT', 'SS'] and required_capacity > 70:
        # For classes with more than 70 students, try to use SEATER_120 rooms first
        seater_120_rooms = {rid: room for rid, room in rooms.items() 
                           if 'SEATER_120' in room['type'].upper()}
        
        # For classes with more than 120 students, use SEATER_240 rooms
        if required_capacity > 120:
            seater_240_rooms = {rid: room for rid, room in rooms.items() 
                              if 'SEATER_240' in room['type'].upper()}
            
            # Try allocating from SEATER_240 first
            room_id = try_room_allocation(seater_240_rooms, 'LEC', required_capacity,
                                        day, start_slot, duration, used_room_ids)
            if room_id:
                return room_id
                
        # Then try SEATER_120 rooms
        room_id = try_room_allocation(seater_120_rooms, 'LEC', required_capacity,
                                    day, start_slot, duration, used_room_ids)
        if room_id:
            return room_id

    # Special handling for labs to get adjacent rooms if needed
    if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
        # Check if student count exceeds standard lab capacity
        if required_capacity > 35:  # Standard lab capacity
            # Try to find adjacent lab rooms
            for room_id, room in rooms.items():
                if room_id in used_room_ids or room['type'].upper() != course_type:
                    continue
                    
                # Check if this room is available
                slots_free = True
                for i in range(duration):
                    if start_slot + i in room['schedule'][day]:
                        slots_free = False
                        break
                
                if slots_free:
                    # Try to find an adjacent room
                    adjacent_room = find_adjacent_lab_room(room_id, rooms)
                    if adjacent_room and adjacent_room not in used_room_ids:
                        # Check if adjacent room is also available
                        adjacent_free = True
                        for i in range(duration):
                            if start_slot + i in rooms[adjacent_room]['schedule'][day]:
                                adjacent_free = False
                                break
                        
                        if adjacent_free:
                            # Mark both rooms as used
                            for i in range(duration):
                                room['schedule'][day].add(start_slot + i)
                                rooms[adjacent_room]['schedule'][day].add(start_slot + i)
                            return f"{room_id},{adjacent_room}"  # Return both room IDs
                            
        # If we don't need two rooms or couldn't find adjacent ones, use regular allocation
        return try_room_allocation(rooms, course_type, required_capacity, day, start_slot, duration, used_room_ids)

    # For lectures and basket courses, try different room types in priority order
    if course_type in ['LEC', 'TUT', 'SS'] or is_basket:
        # First try regular lecture rooms
        lecture_rooms = {rid: room for rid, room in rooms.items() 
                        if 'LECTURE_ROOM' in room['type'].upper()}
        
        # For basket courses, need special room allocation
        if is_basket:
            basket_group = get_basket_group(course_code)
            basket_used_rooms = set()
            basket_group_rooms = {}  # Track rooms already allocated to this basket group
            
            # Track room usage count
            room_usage = {rid: sum(len(room['schedule'][d]) for d in range(len(DAYS))) 
                         for rid, room in rooms.items()}
            
            # Sort lecture rooms by usage count
            sorted_lecture_rooms = dict(sorted(lecture_rooms.items(), 
                                             key=lambda x: room_usage[x[0]]))
            
            # Check room availability for the sorted rooms
            for room_id, room in sorted_lecture_rooms.items():
                is_used = False
                for slot in range(start_slot, start_slot + duration):
                    if slot in rooms[room_id]['schedule'][day]:
                        # Check if room is used by any course from same basket group
                        if slot in timetable[day]:
                            slot_data = timetable[day][slot]
                            if (slot_data['classroom'] == room_id and 
                                slot_data['type'] is not None):
                                slot_code = slot_data.get('code', '')
                                if get_basket_group(slot_code) == basket_group:
                                    basket_group_rooms[slot_code] = room_id
                                else:
                                    basket_used_rooms.add(room_id)
                        is_used = True
                        break
                
                # Room is free for this time slot
                if not is_used and room_id not in basket_used_rooms:
                    if 'capacity' in room and room['capacity'] >= required_capacity:
                        # Mark slots as used
                        for i in range(duration):
                            room['schedule'][day].add(start_slot + i)
                        return room_id
            
            # If no unused room found, try existing basket group rooms
            if course_code in basket_group_rooms:
                return basket_group_rooms[course_code]
            
            # Try remaining rooms through regular allocation
            room_id = try_room_allocation(lecture_rooms, 'LEC', required_capacity,
                                        day, start_slot, duration, basket_used_rooms)
            
            if room_id:
                basket_group_rooms[course_code] = room_id
            
            return room_id

        # For non-basket courses, use regular lecture rooms
        return try_room_allocation(lecture_rooms, 'LEC', required_capacity,
                                 day, start_slot, duration, used_room_ids)
    
    # For labs, use existing logic
    return try_room_allocation(rooms, course_type, required_capacity,
                             day, start_slot, duration, used_room_ids)

def check_faculty_daily_components(professor_schedule, faculty, day, department, semester, section, timetable, course_code=None, activity_type=None):
    """Check faculty/course scheduling constraints for the day"""
    component_count = 0
    faculty_courses = set()  # Track faculty's courses 
    
    # Check all slots for this day
    for slot in timetable[day].values():
        if slot['faculty'] == faculty and slot['type'] in ['LEC', 'LAB', 'TUT']:
            slot_code = slot.get('code', '')
            if slot_code:
                # For non-basket courses
                if not is_basket_course(slot_code):
                    component_count += 1
                # For basket courses, only count if not already counted
                elif slot_code not in faculty_courses:
                    component_count += 1
                    faculty_courses.add(slot_code)
                    
    # Special handling for basket courses - allow parallel scheduling
    if course_code and is_basket_course(course_code):
        basket_group = get_basket_group(course_code)
        existing_slots = get_basket_group_slots(timetable, day, basket_group)
        if existing_slots:
            # For basket courses, check only non-basket components
            return component_count < 3  # Allow more flexibility for basket courses
    
    return component_count < 2  # Keep max 2 components per day limit for regular courses

def check_faculty_course_gap(professor_schedule, timetable, faculty, course_code, day, start_slot):
    """Check if there is sufficient gap (3 hours) between sessions of same course"""
    min_gap_hours = 3
    slots_per_hour = 2  # Assuming 30-min slots
    required_gap = min_gap_hours * slots_per_hour
    
    # Check previous slots
    for i in range(max(0, start_slot - required_gap), start_slot):
        if i in professor_schedule[faculty][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
                
    # Check next slots  
    for i in range(start_slot + 1, min(len(TIME_SLOTS), start_slot + required_gap)):
        if i in professor_schedule[faculty][day]:
            slot_data = timetable[day][i]
            if slot_data['code'] == course_code and slot_data['type'] in ['LEC', 'TUT']:
                return False
    
    return True

def is_preferred_slot(faculty, day, time_slot, faculty_preferences):
    """Check if a time slot is within faculty's preferences"""
    if faculty not in faculty_preferences:
        return True  # No preferences specified, any slot is fine
        
    prefs = faculty_preferences[faculty]
    
    # Check if day is preferred
    if prefs['preferred_days'] and DAYS[day] not in prefs['preferred_days']:
        return False
        
    # Check if time is within preferred ranges
    if prefs['preferred_times']:
        slot_start, slot_end = time_slot
        for pref_start, pref_end in prefs['preferred_times']:
            if (slot_start >= pref_start and slot_end <= pref_end):
                return True
        return False
        
    return True  # No time preferences specified

def is_lecture_scheduled(timetable, day, start_slot, end_slot):
    """Check if there's a lecture scheduled in the given time range"""
    for slot in range(start_slot, end_slot):
        if (slot < len(timetable[day]) and 
            timetable[day][slot]['type'] and 
            timetable[day][slot]['type'] in ['LEC', 'LAB', 'TUT']):
            return True
    return False

def get_best_slots(timetable, professor_schedule, faculty, day, duration, semester, department):
    """Find best available consecutive slots in a day"""
    best_slots = []
    
    for start_slot in range(len(TIME_SLOTS) - duration + 1):
        slots_free = True
        # Check each slot in the duration
        for i in range(duration):
            current_slot = start_slot + i
            # Different handling for LAB vs other activities
            if duration == LAB_DURATION:
                # For labs, block slots even if they have basket courses
                # This ensures labs get priority over basket courses
                if (current_slot in professor_schedule[faculty][day] or
                    timetable[day][current_slot]['type'] is not None or  # Block any existing schedule
                    is_break_time(TIME_SLOTS[current_slot], semester)):
                    slots_free = False
                    break
            else:
                # Original logic for lectures/tutorials
                if (current_slot in professor_schedule[faculty][day] or
                    (timetable[day][current_slot]['type'] is not None and
                     not is_basket_course(timetable[day][current_slot].get('code', ''))) or
                    is_break_time(TIME_SLOTS[current_slot], semester)):
                    slots_free = False
                    break

        if slots_free:
            # Prioritize morning slots (before lunch) for labs
            if duration == LAB_DURATION:
                slot_time = TIME_SLOTS[start_slot][0]
                if slot_time < time(12, 30):  # Before lunch
                    best_slots.insert(0, start_slot)  # Add to beginning of list (higher priority)
                else:
                    best_slots.append(start_slot)
            else:
                best_slots.append(start_slot)
    
    return best_slots

def is_slot_reserved(slot, day, semester, department, reserved_slots):
    """Check if a time slot is reserved for this semester and department"""
    if day not in reserved_slots:
        return False
        
    slot_start, slot_end = slot
    
    # Check each reservation
    for (dept, semesters), slots in reserved_slots[day].items():
        # Match if department is ALL or matches exactly
        if dept == 'ALL' or dept == department:
            # Match if semester is in the expanded semester list
            if str(semester) in semesters or any(str(semester).startswith(s) for s in semesters):
                for reserved_start, reserved_end in slots:
                    if (slot_start >= reserved_start and slot_start < reserved_end) or \
                       (slot_end > reserved_start and slot_end <= reserved_end):
                        return True
    return False

def get_course_priority(course):
    """Calculate course scheduling priority based on constraints"""
    priority = 0
    code = str(course['Course Code'])
    
    # Give regular course labs highest priority with much higher weight
    if pd.notna(course['P']) and course['P'] > 0 and not is_basket_course(code):
        priority += 10  # Increased from 5 to 10 for regular labs
        if 'CS' in code or 'EC' in code:  # Extra priority for CS/EC labs
            priority += 2
    elif is_basket_course(code):
        priority += 1  # Keep lowest priority for basket courses
    elif pd.notna(course['L']) and course['L'] > 2:
        priority += 3  # Regular lectures priority
    elif pd.notna(course['T']) and course['T'] > 0:
        priority += 2  # Tutorial priority
    return priority

def calculate_required_slots(course):
    """Calculate how many slots needed based on L, T, P, S values and credits"""
    l = float(course['L']) if pd.notna(course['L']) else 0  # Lecture credits
    t = int(course['T']) if pd.notna(course['T']) else 0    # Tutorial hours
    p = int(course['P']) if pd.notna(course['P']) else 0    # Lab hours
    s = int(course['S']) if pd.notna(course['S']) else 0    # Self study hours
    c = int(course['C']) if pd.notna(course['C']) else 0    # Total credits
    
    # Check if course is self-study only
    if s > 0 and l == 0 and t == 0 and p == 0:
        return 0, 0, 0, 0
        
    # Calculate number of lecture sessions based on credits
    lecture_sessions = 0
    if l > 0:
        # For 3 credits = 2 sessions of 1.5 hours each
        # For 2 credits = 1 session of 1.5 hours plus a 1 hour session
        # For 1 credit = 1 session of 1.5 hours
        lecture_sessions = max(1, round(l * 2/3))  # Scale credits to sessions
    
    # Other calculations remain the same
    tutorial_sessions = t  
    lab_sessions = p // 2  # 2 hours per lab session
    self_study_sessions = s // 4 if (l > 0 or t > 0 or p > 0) else 0
    
    return lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions

def select_faculty(faculty_str):
    """Select a faculty from potentially multiple options."""
    if '/' in faculty_str:
        # Split by slash and strip whitespace
        faculty_options = [f.strip() for f in faculty_str.split('/')]
        return faculty_options[0]  # Take first faculty as default
    return faculty_str

class UnscheduledComponent:
    def __init__(self, department, semester, code, name, faculty, component_type, sessions, section='', reason=''):
        self.department = department
        self.semester = semester
        self.code = code
        self.name = name
        self.faculty = faculty 
        self.component_type = component_type
        self.sessions = sessions
        self.section = section
        self.reason = reason
        
    def __eq__(self, other):
        if not isinstance(other, UnscheduledComponent):
            return False
        return (self.department == other.department and
                self.semester == other.semester and
                self.code == other.code and
                self.component_type == other.component_type and
                self.section == other.section)
    
    def __hash__(self):
        return hash((self.department, self.semester, self.code, self.component_type, self.section))

def unscheduled_reason(course, department, semester, professor_schedule, rooms, component_type, check_attempts):
    """Generate detailed reason why a course component couldn't be scheduled"""
    faculty = course['Faculty']
    code = str(course['Course Code'])
    
    # Check faculty availability
    faculty_slots_used = 0
    for day in range(len(DAYS)):
        if faculty in professor_schedule and day in professor_schedule[faculty]:
            faculty_slots_used += len(professor_schedule[faculty][day])
    
    # If faculty is heavily scheduled
    if faculty_slots_used > 20:  # Threshold: 10 hours of teaching per week
        return f"Faculty '{faculty}' already has {faculty_slots_used/2:.1f} hours of teaching scheduled"
    
    # Check room availability issues
    if component_type == 'LAB':
        lab_rooms_available = False
        for _, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                lab_rooms_available = True
                break
        
        if not lab_rooms_available:
            return "No suitable lab rooms available in the system"
        
        # Check if room is overbooked
        lab_rooms_free_slots = 0
        for rid, room in rooms.items():
            if 'LAB' in room['type'].upper() or 'COMPUTER' in room['type'].upper():
                total_slots = len(DAYS) * (len(TIME_SLOTS) - LAB_DURATION)
                used_slots = sum(len(room['schedule'].get(day, [])) for day in range(len(DAYS)))
                lab_rooms_free_slots += (total_slots - used_slots)
        
        if lab_rooms_free_slots < 5:  # Very few lab slots left
            return f"Lab rooms almost fully booked ({lab_rooms_free_slots} slots left)"
    
    # Check for large classes with insufficient large rooms
    if 'total_students' in course and pd.notna(course['total_students']):
        try:
            total_students = int(course['total_students'])
            if total_students > 100:
                large_rooms_available = False
                for _, room in rooms.items():
                    if room['type'].upper() == 'SEATER_120' or room['type'].upper() == 'SEATER_240':
                        large_rooms_available = True
                        break
                
                if not large_rooms_available:
                    return f"No rooms available with capacity for {total_students} students"
        except (ValueError, TypeError):
            pass
    
    # Check timeslot conflicts with other courses in same department/semester
    if check_attempts > 800:  # If we made many attempts but still couldn't find a slot
        return f"No suitable timeslot found after {check_attempts} attempts - heavy scheduling conflicts"
        
    # Default reason
    duration_map = {
        'LEC': f"{LECTURE_DURATION/2} hour",
        'LAB': f"{LAB_DURATION/2} hour",
        'TUT': f"{TUTORIAL_DURATION/2} hour"
    }
    duration_str = duration_map.get(component_type, "")
    
    return f"Could not find compatible {duration_str} timeslot for {code} {component_type} with faculty {faculty}"

if __name__ == "__main__":
    generate_all_timetables()
    check_unscheduled_courses()
    generate_faculty_timetables()